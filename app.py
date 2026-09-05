from __future__ import annotations

from io import BytesIO
from pathlib import Path
from datetime import datetime, timezone, timedelta
import hashlib
import tempfile
import re
import gc
import shutil
import uuid

import pandas as pd
import numpy as np
import streamlit as st
import matplotlib.pyplot as plt
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from plotly.colors import qualitative
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from procesador import (
    procesar_archivo,
    clasificar_tipo_disparo_v33,
    generar_grafico,
    generar_plano_zda_png,
    extraer_plano_navegacion_png,
)


# ==========================================================
# CONFIGURACIÓN
# ==========================================================

APP_VERSION_INTERNAL = "V34.96-PYTHON-CUCHULA-NEGRO"
PUBLIC_VERSION = "v1.0"
CACHE_SCHEMA_VERSION = "v34_44_python_masivo_150_zda_20260826"
TIPOS_DISPARO = ["FRENTE", "SELLADA", "ESTOCADA Y/O CORRECCIONES"]
COLORES = qualitative.Plotly

st.set_page_config(page_title=f"EBR Drill Analytics · Piloto {PUBLIC_VERSION}", page_icon="⛏️", layout="wide")
st.title("EBR Drill Analytics")
st.caption(f"Piloto {PUBLIC_VERSION} · Análisis de reportes de perforación de equipos Jumbo Sandvik")
st.info(
    "Consolida y analiza información de perforación de reportes PDF y ZDA, mostrando automatización por jumbo y brazo, "
    "longitud perforada en barrenos Cut, tiempos de ciclo, clasificación de disparos y exportación de datos a Excel."
)

st.markdown(
    """
    <style>
    /* Reserva espacio debajo de la barra superior de Streamlit
       para evitar que el título quede recortado/oculto por el header. */
    [data-testid="stAppViewBlockContainer"],
    .block-container {
        padding-top: 3.25rem !important;
        padding-bottom: 6rem !important;
    }

    h1 {
        font-size: 1.75rem !important;
        line-height: 1.20 !important;
        margin-top: 0 !important;
        margin-bottom: 0.15rem !important;
        padding-top: 0.15rem !important;
        overflow: visible !important;
    }
    h2 { font-size: 1.22rem !important; margin-top: 0.7rem !important; }
    h3 { font-size: 1.00rem !important; margin-top: 0.55rem !important; margin-bottom: 0.25rem !important; }
    [data-testid="stMetricLabel"] { font-size: 0.75rem !important; }
    [data-testid="stMetricValue"] { font-size: 1.20rem !important; }
    [data-testid="stDataFrame"] div[role="grid"] { font-size: 0.76rem !important; }
    [data-testid="stExpander"] summary { font-size: 0.88rem !important; font-weight: 600 !important; }
    .stDownloadButton button, .stButton button { min-height: 2.15rem !important; font-size: 0.80rem !important; }
    </style>
    """,
    unsafe_allow_html=True,
)


# ==========================================================
# ESTADO
# ==========================================================

if st.session_state.get("cache_schema_version") != CACHE_SCHEMA_VERSION:
    st.session_state.procesados = {}
    st.session_state.uploader_version = st.session_state.get("uploader_version", 0) + 1
    st.session_state.cache_schema_version = CACHE_SCHEMA_VERSION

if "uploader_version" not in st.session_state:
    st.session_state.uploader_version = 0
if "procesados" not in st.session_state:
    st.session_state.procesados = {}
if "session_disk_id" not in st.session_state:
    st.session_state.session_disk_id = uuid.uuid4().hex
if "staged_queue" not in st.session_state:
    st.session_state.staged_queue = []
if "auto_process_staged" not in st.session_state:
    st.session_state.auto_process_staged = False


def _session_work_dir() -> Path:
    root = Path(tempfile.gettempdir()) / "ebr_drill_massive"
    path = root / st.session_state.session_disk_id
    path.mkdir(parents=True, exist_ok=True)
    (path / "uploads").mkdir(exist_ok=True)
    (path / "visuals").mkdir(exist_ok=True)
    return path


# Streamlit vuelve a ejecutar el script completo ante cualquier cambio de widget.
# st.fragment permite que filtros y checkboxes vuelvan a ejecutar SOLO su bloque,
# evitando recalcular Excel, otros gráficos y resultados individuales.
if hasattr(st, "fragment"):
    fragment = st.fragment
else:
    def fragment(func):
        return func


def limpiar_analisis():
    work_dir = None
    try:
        work_dir = _session_work_dir()
    except Exception:
        pass

    st.session_state.procesados = {}
    st.session_state.staged_queue = []
    st.session_state.auto_process_staged = False
    st.session_state.uploader_version += 1

    if work_dir and work_dir.exists():
        shutil.rmtree(work_dir, ignore_errors=True)

    # Nuevo directorio limpio para la misma sesión.
    st.session_state.session_disk_id = uuid.uuid4().hex

    for key in list(st.session_state.keys()):
        if key.startswith((
            "global_",
            "_global_",
            "opt_",
            "auto_",
            "cut_",
            "zda_",
            "class_",
            "lbl_",
            "desglosar_",
            "detalle_ciclo_",
        )):
            del st.session_state[key]


def tipo_roca_desde_plan_texto(plan_perforacion):
    """
    Extrae el tipo/clase de roca desde Plan_Perforacion / drill_plan.
    Ejemplos:
      "MALLA E.E. 4.5x4.5 III-B" -> "III-B"
      "MALLA E.B. 5.0x5.0 IV-A"   -> "IV-A"

    Se acepta cualquier número romano seguido de guion y código
    alfanumérico para no limitar futuros valores.
    """
    texto = str(plan_perforacion or "").upper()
    m = re.search(
        r"\b([IVXLCDM]+)\s*[-–—]\s*([A-Z0-9]+)\b",
        texto,
        re.IGNORECASE,
    )
    if not m:
        return "SIN DATO"

    return f"{m.group(1).upper()}-{m.group(2).upper()}"


# ==========================================================
# FILTROS Y PARÁMETROS GLOBALES
# ==========================================================

def _valores_detectados_desde_cache():
    """
    Obtiene jumbos, tipos de disparo y tipos de roca desde los
    resultados ya procesados. No presupone cuántos equipos ni
    cuántas clases de roca existen.
    """
    jumbos = []
    tipos = []
    rocas = []
    operadores = []

    for r in st.session_state.procesados.values():
        if r.get("error"):
            continue

        rep = r.get("resumen_reporte") or {}
        jumbo = rep.get("Jumbo")
        tipo = rep.get("Tipo_Disparo")
        roca = tipo_roca_desde_plan_texto(
            rep.get("Plan_Perforacion")
        )
        operador = (
            rep.get("Operador_ZDA")
            or rep.get("Operador")
            or None
        )

        if jumbo and str(jumbo).strip():
            jumbos.append(str(jumbo).strip())
        if tipo and str(tipo).strip():
            tipos.append(str(tipo).strip())
        if roca and str(roca).strip():
            rocas.append(str(roca).strip())
        if operador and str(operador).strip():
            operadores.append(str(operador).strip())

    # Orden natural para JUMB001, JUMB002... y luego series/otros.
    def jumbo_sort_key(valor):
        m = re.fullmatch(r"JUMB(\d+)", str(valor), re.IGNORECASE)
        if m:
            return (0, int(m.group(1)))
        return (1, str(valor))

    jumbos = sorted(set(jumbos), key=jumbo_sort_key)
    tipos_presentes = set(tipos)
    tipos = [t for t in TIPOS_DISPARO if t in tipos_presentes]

    # Si apareciera una clasificación futura no contemplada, no se pierde.
    tipos_extra = sorted(tipos_presentes - set(TIPOS_DISPARO))
    tipos.extend(tipos_extra)

    # SIN DATO al final; los demás valores alfabéticamente.
    rocas_presentes = set(rocas)
    rocas = sorted(r for r in rocas_presentes if r != "SIN DATO")
    if "SIN DATO" in rocas_presentes:
        rocas.append("SIN DATO")

    operadores = sorted(set(operadores))

    return jumbos, tipos, rocas, operadores


def _sincronizar_multiselect_dinamico(key, options, previous_options_key):
    """
    Mantiene la selección del usuario y agrega automáticamente
    cualquier opción nueva detectada tras procesar más archivos.
    """
    options = list(options)
    prev_options = list(st.session_state.get(previous_options_key, []))

    if key not in st.session_state:
        st.session_state[key] = options.copy()
    else:
        actual = [
            x for x in st.session_state.get(key, [])
            if x in options
        ]
        nuevos = [
            x for x in options
            if x not in prev_options
        ]
        for x in nuevos:
            if x not in actual:
                actual.append(x)
        st.session_state[key] = actual

    st.session_state[previous_options_key] = options.copy()


jumbos_detectados, tipos_detectados, rocas_detectadas, operadores_detectados = _valores_detectados_desde_cache()

# Variables siempre definidas aunque aún no existan datos.
global_jumbos = []
global_tipos = []
global_rocas = []
global_operadores = []
global_lbl_auto = False
global_line_auto = False
global_lbl_arm = False
global_lbl_cut = False
global_lbl_zda = False
sidebar_fecha_container = None

with st.sidebar:
    st.header("Filtros y parámetros")
    sidebar_fecha_container = st.container()

    if not jumbos_detectados:
        st.info(
            "Los filtros se habilitarán automáticamente después de procesar "
            "los primeros archivos PDF/ZDA."
        )
    else:
        _sincronizar_multiselect_dinamico(
            "global_jumbos",
            jumbos_detectados,
            "_global_jumbos_options_prev",
        )
        _sincronizar_multiselect_dinamico(
            "global_tipos",
            tipos_detectados,
            "_global_tipos_options_prev",
        )
        _sincronizar_multiselect_dinamico(
            "global_rocas",
            rocas_detectadas,
            "_global_rocas_options_prev",
        )
        _sincronizar_multiselect_dinamico(
            "global_operadores",
            operadores_detectados,
            "_global_operadores_options_prev",
        )

        global_jumbos = st.multiselect(
            "Jumbos",
            jumbos_detectados,
            key="global_jumbos",
        )

        global_tipos = st.multiselect(
            "Tipo de disparo",
            tipos_detectados,
            key="global_tipos",
        )

        global_rocas = st.multiselect(
            "Tipo de roca",
            rocas_detectadas,
            key="global_rocas",
        )

        global_operadores = st.multiselect(
            "Operadores",
            operadores_detectados,
            key="global_operadores",
            help=(
                "Se muestran únicamente los operadores detectados "
                "en los archivos ZDA procesados."
            ),
        )

        st.caption(
            "Los equipos se detectan automáticamente a partir de los archivos procesados."
        )
        st.caption(
            "Estos filtros se aplican a los gráficos y resúmenes consolidados."
        )
        st.caption(
            "**Clasificación:** FRENTE > 45 barrenos · SELLADA 25–45 barrenos · "
            "ESTOCADA Y/O CORRECCIONES < 25 barrenos. Conteo sobre barrenos de frente "
            "(Bottom + Easer + Cut + Contour); Reaming y Casing no se consideran."
        )

        st.divider()
        st.subheader("Opciones de gráficos")

        global_lbl_auto = st.checkbox(
            "Etiquetas · movimiento automático",
            value=False,
            key="opt_lbl_auto",
        )
        global_line_auto = st.checkbox(
            "Línea curva · movimiento automático",
            value=True,
            key="opt_line_auto",
            help=(
                "Desmarcado: muestra solo los puntos. "
                "Marcado: agrega una curva suavizada que conecta los ciclos sin modificar los valores reales de los puntos."
            ),
        )
        global_lbl_arm = st.checkbox(
            "Etiquetas · uso por brazo",
            value=False,
            key="opt_lbl_arm",
        )
        global_lbl_cut = st.checkbox(
            "Etiquetas · barrenos Cut",
            value=False,
            key="opt_lbl_cut",
        )
        global_lbl_zda = st.checkbox(
            "Etiquetas · tiempos de ciclo",
            value=False,
            key="opt_lbl_zda",
        )

        st.caption(
            "Los filtros no eliminan datos del Excel exportado; la exportación conserva "
            "todos los archivos procesados."
        )


# ==========================================================
# HELPERS
# ==========================================================

def seccion_desde_plan_texto(plan_perforacion):
    """
    Devuelve una etiqueta corta de sección desde Plan_Perforacion.
    Ej.: "MALLA E.E. 4.5x4.5 III-B" -> "4.5 x 4.5"
    """
    texto = str(plan_perforacion or "")
    m = re.search(
        r"(\d+(?:[.,]\d+)?)\s*[xX×]\s*(\d+(?:[.,]\d+)?)",
        texto,
    )
    if not m:
        return "-"

    a = float(m.group(1).replace(",", "."))
    b = float(m.group(2).replace(",", "."))
    return f"{a:.1f} x {b:.1f}"




def hash_archivo(uploaded_file) -> str:
    # Evita uploaded_file.getvalue(), que crea una copia completa del archivo
    # en RAM. getbuffer() entrega una vista de memoria sin duplicar el ZDA/PDF.
    h = hashlib.sha256()
    h.update(uploaded_file.getbuffer())
    h.update(CACHE_SCHEMA_VERSION.encode("utf-8"))
    return h.hexdigest()


def hash_archivo_en_disco(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    h.update(CACHE_SCHEMA_VERSION.encode("utf-8"))
    return h.hexdigest()


def preparar_archivos_en_disco(uploaded_files):
    """
    Primera fase del modo masivo.

    Copia los UploadedFile a disco temporal SIN procesarlos todavía.
    Después se fuerza un rerun para que Streamlit libere del uploader los
    bytes de los 100–150 ZDA antes de comenzar el parsing intensivo.
    """
    uploads_dir = _session_work_dir() / "uploads"
    nuevos = []

    for n, archivo in enumerate(uploaded_files, start=1):
        suffix = Path(archivo.name).suffix.lower()
        safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(archivo.name).name)
        disk_path = uploads_dir / f"{uuid.uuid4().hex[:10]}_{safe_name}"

        archivo.seek(0)
        with disk_path.open("wb") as out:
            shutil.copyfileobj(archivo, out, length=1024 * 1024)

        clave = hash_archivo_en_disco(disk_path)
        if clave in st.session_state.procesados:
            disk_path.unlink(missing_ok=True)
            continue

        # Evitar duplicar el mismo archivo en una cola ya preparada.
        if any(x.get("clave") == clave for x in st.session_state.staged_queue):
            disk_path.unlink(missing_ok=True)
            continue

        item = {
            "clave": clave,
            "nombre": archivo.name,
            "path": str(disk_path),
            "size": disk_path.stat().st_size,
        }
        st.session_state.staged_queue.append(item)
        nuevos.append(item)

    return nuevos


def fmt(valor, dec=1, sufijo=""):
    if valor is None or pd.isna(valor):
        return "-"
    return f"{float(valor):.{dec}f}{sufijo}"


def figura_a_png(fig) -> bytes:
    buffer = BytesIO()
    fig.savefig(buffer, format="png", dpi=250, bbox_inches="tight")
    buffer.seek(0)
    return buffer.getvalue()


def concatenar_dataframes(resultados, key, solo_ok=False):
    dfs = []
    for r in resultados:
        if solo_ok and str(r.get("resumen_reporte", {}).get("Estado")) != "OK":
            continue
        df = r.get(key)
        if isinstance(df, pd.DataFrame) and not df.empty:
            dfs.append(df)
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()


def asegurar_fechahora(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["FechaHora"] = pd.to_datetime(
        out.get("Fecha_Inicio", pd.Series(index=out.index, dtype=object)).fillna("")
        + " "
        + out.get("Hora_Inicio", pd.Series(index=out.index, dtype=object)).fillna(""),
        format="%d/%m/%Y %H:%M:%S",
        errors="coerce",
    )
    faltan = out["FechaHora"].isna()
    if faltan.any() and "Fecha_Inicio" in out.columns:
        out.loc[faltan, "FechaHora"] = pd.to_datetime(
            out.loc[faltan, "Fecha_Inicio"], format="%d/%m/%Y", errors="coerce"
        )
    return out[out["FechaHora"].notna()].sort_values(["FechaHora", "Jumbo", "Ciclo"])


def smart_annotations(points, x_window_hours=18, y_window=3.0, font_size=11):
    """Equivalente del buildSmartAnnotations de la HTML V33."""
    out, used = [], []
    points = sorted(points, key=lambda p: (pd.Timestamp(p["x"]), float(p["y"])))
    x_window = pd.Timedelta(hours=x_window_hours)
    for i, p in enumerate(points):
        base_sign = -1 if int(p.get("rank", i)) % 2 == 0 else 1
        for level in range(8):
            yshift = base_sign * (18 + level * 14)
            clash = False
            for u in used:
                if (
                    abs(pd.Timestamp(u["x"]) - pd.Timestamp(p["x"])) <= x_window
                    and abs(float(u["y"]) - float(p["y"])) <= y_window
                    and abs(u["yshift"] - yshift) < 14
                ):
                    clash = True
                    break
            if not clash:
                used.append({"x": p["x"], "y": p["y"], "yshift": yshift})
                out.append(
                    dict(
                        x=p["x"], y=p["y"], xref="x", yref="y",
                        showarrow=False, yshift=yshift, text=p["text"],
                        align="center", font=dict(size=font_size, color="#334155"),
                        bgcolor="rgba(255,255,255,0.88)",
                        bordercolor="rgba(203,213,225,0.95)", borderpad=3,
                    )
                )
                break
    return out


def base_layout(height=450, **kwargs):
    layout = dict(
        height=height,
        paper_bgcolor="#ffffff",
        plot_bgcolor="#ffffff",
        font=dict(family="Arial, sans-serif", size=12, color="#334155"),
        margin=dict(l=70, r=30, t=70, b=70),
        hovermode="closest",
        legend=dict(orientation="h", y=-0.18, x=0),
        xaxis=dict(showgrid=True, gridcolor="#eef2f7", zeroline=False),
        yaxis=dict(showgrid=True, gridcolor="#eef2f7", zeroline=False),
    )
    layout.update(kwargs)
    return layout



# ==========================================================
# EXCEL / BD-PERFO - PUBLICACIÓN V1.0
# ==========================================================

BD_PERFO_COLUMNS = [
    "MES","SEMANA","AÑO","FECHA","TURNO","JEFE DE TURNO","OPERADOR","JUMBO",
    "NIVEL","BLOCK","LABOR","TIPO DISPARO","SECCIÓN","RMR","TIP. DE ROCA",
    "TIPO EXP.","Hora 1° taladro","Tiempo de perforación efec.","Tiempo de perforación",
    "Tiempo de movimiento ","T. mov. Manual","T.mov. Autom.","T. AUTOMÁTICO",
    "# TALADROS","T. MANUAL","Av. Perf","Av.Top","Av.Real","Eficiencia",
    "Obj. Automático","CONLABOR","Tipo Mat","Ciclo","%AutoDrill","% Auto B1",
    "% Auto B2","Mediana Cut (m)","Promedio Cut (m)","Hora 1° martillo"
]

BD_JUMBO_ALIAS = {"JUMB001": "JF01", "JUMB002": "JF02"}
BD_MESES_OPERATIVOS = [
    "Enero","Febrero","Marzo","Abril","Mayo","Junio",
    "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre",
]


def _bd_num(v):
    if v is None or pd.isna(v):
        return None
    try:
        n = float(v)
        return n if pd.notna(n) else None
    except Exception:
        return None


def _bd_parse_datetime(row):
    candidatos = []
    inicio_perf = row.get("Inicio_Perforacion")
    if inicio_perf is not None and not pd.isna(inicio_perf):
        candidatos.append(str(inicio_perf))
    fecha = row.get("Fecha_Inicio")
    hora = row.get("Hora_Inicio")
    if fecha is not None and not pd.isna(fecha):
        candidatos.append(
            f"{fecha} {hora if hora is not None and not pd.isna(hora) else '00:00:00'}"
        )

    for s in candidatos:
        dt = pd.to_datetime(s, format="%d/%m/%Y %H:%M:%S", errors="coerce")
        if pd.notna(dt):
            return dt.to_pydatetime()
    return None


def _bd_mes_operativo(dt):
    """
    Mes operativo para la hoja BD-PERFO.

    Regla:
    - Del día 01 al 25 -> pertenece al mes calendario actual.
    - Del día 26 al último día del mes -> pertenece al mes operativo siguiente.

    Ejemplos:
    - 26/07 al 25/08 = Agosto
    - 26/08 al 25/09 = Septiembre
    - 26/12 al 25/01 = Enero
    """
    if dt is None:
        return None

    mes_operativo = dt.month + (1 if dt.day >= 26 else 0)
    if mes_operativo > 12:
        mes_operativo = 1

    return BD_MESES_OPERATIVOS[mes_operativo - 1]


def _bd_turno(dt):
    if dt is None:
        return None
    return "DIA" if 7 <= dt.hour < 19 else "NOCHE"


def _bd_hora_excel(dt):
    if dt is None:
        return None
    return (dt.hour * 3600 + dt.minute * 60 + dt.second) / 86400.0


def _bd_plan_fields(plan):
    p = "" if plan is None or pd.isna(plan) else str(plan)
    seccion = roca = tipo_exp = None

    m = re.search(r"(\d+(?:[.,]\d+)?)\s*[xX×]\s*(\d+(?:[.,]\d+)?)", p)
    if m:
        seccion = f"{m.group(1).replace(',', '.')}x{m.group(2).replace(',', '.')}"

    m = re.search(r"\b(VI|IV|V|III|II|I)\s*-\s*([A-C])\b", p, re.IGNORECASE)
    if m:
        roca = f"{m.group(1).upper()}-{m.group(2).upper()}"

    if re.search(r"\bE\s*\.\s*E\s*\.?", p, re.IGNORECASE):
        tipo_exp = "Encartuchada"
    elif re.search(r"\bE\s*\.\s*B\s*\.?", p, re.IGNORECASE):
        tipo_exp = "Bombeable"

    return seccion, roca, tipo_exp


def _bd_labor_fields(row):
    curve_raw = row.get("Tabla_Curvas")
    tunnel_raw = row.get("Labor")
    curve = "" if curve_raw is None or pd.isna(curve_raw) else str(curve_raw).strip()
    tunnel = "" if tunnel_raw is None or pd.isna(tunnel_raw) else str(tunnel_raw).strip()
    src = curve or tunnel
    nivel = block = labor = None

    if src:
        m = re.search(r"\bNV\s*:?\s*(\d+)\b", src, re.IGNORECASE)
        if m:
            nivel = int(m.group(1))

        m = re.search(r"\b(?:BLOCK|B)\s*:?\s*(\d+)\b", src, re.IGNORECASE)
        if m:
            block = int(m.group(1))

        if curve:
            for token in [t for t in re.split(r"\s+", curve) if t]:
                if re.match(r"^(?:NV|B|BLOCK)\s*:?\d+$", token, re.IGNORECASE):
                    continue
                if re.match(r"^\d+$", token):
                    continue
                labor = token.upper()
                break

        if not labor and tunnel:
            for key, value in re.findall(
                r"\b([A-Za-z]{1,8})\s*:\s*([A-Za-z0-9._-]+)", tunnel
            ):
                if key.upper() not in {"NV","OP","T","TURN","RMR"}:
                    labor = f"{key}{value}".upper()
                    break

    return nivel, block, labor


def _bd_operator_map(df_reportes):
    out = {}
    if df_reportes.empty or "Fuente" not in df_reportes.columns:
        return out

    pdfs = df_reportes[
        df_reportes["Fuente"].astype(str).str.upper().eq("PDF")
    ].copy()

    for _, r in pdfs.iterrows():
        op = r.get("Operario")
        if op is None or pd.isna(op) or str(op).strip() == "":
            continue
        op = str(op).strip()
        serie_base = re.sub(
            r"-(?:\d+|L)$",
            "",
            str(r.get("Numero_Serie") or ""),
            flags=re.IGNORECASE,
        )
        keys = [
            f"{r.get('Jumbo','')}|{r.get('Ciclo','')}",
            f"{serie_base}|{r.get('Ciclo','')}",
        ]
        for key in keys:
            if key.startswith("|") or key.endswith("|"):
                continue
            out.setdefault(key, [])
            if op not in out[key]:
                out[key].append(op)

    return {k: " / ".join(v) for k, v in out.items()}


def _bd_operador_para_zda(row, op_map):
    serie_base = re.sub(
        r"-(?:\d+|L)$",
        "",
        str(row.get("Numero_Serie") or ""),
        flags=re.IGNORECASE,
    )
    keys = [
        f"{row.get('Jumbo','')}|{row.get('Ciclo','')}",
        f"{serie_base}|{row.get('Ciclo','')}",
    ]
    for key in keys:
        if key in op_map:
            return op_map[key]
    return None


def _bd_operador_exportado(row, op_map):
    """
    Operador para BD-PERFO.

    Prioridad:
    1) Operador_ZDA leído directamente de round.txt (OP:...)
    2) Operador normalizado disponible en el registro
    3) Cruce histórico con PDF por Jumbo/Serie + Ciclo
    4) Vacío
    """
    for campo in ("Operador_ZDA", "Operador"):
        valor = row.get(campo)
        if valor is not None and not pd.isna(valor):
            texto = str(valor).strip()
            if texto:
                return texto

    return _bd_operador_para_zda(row, op_map)


def _bd_cut_stats(df_detalle, jumbo, ciclo):
    if df_detalle.empty or "Longitud_roca_m" not in df_detalle.columns:
        return None, None

    mask = pd.Series(True, index=df_detalle.index)
    if "Fuente" in df_detalle.columns:
        mask &= df_detalle["Fuente"].astype(str).str.upper().eq("ZDA")
    if "Jumbo" in df_detalle.columns:
        mask &= df_detalle["Jumbo"].astype(str).eq(str(jumbo))
    if "Ciclo" in df_detalle.columns:
        mask &= df_detalle["Ciclo"].astype(str).eq(str(ciclo))
    if "Tipo" in df_detalle.columns:
        mask &= df_detalle["Tipo"].astype(str).str.upper().eq("CUT")

    vals = pd.to_numeric(
        df_detalle.loc[mask, "Longitud_roca_m"], errors="coerce"
    ).dropna()

    if vals.empty:
        return None, None
    return float(vals.median()), float(vals.mean())


def _bd_primer_martillo(df_detalle, jumbo, ciclo):
    if df_detalle.empty or "Inicio_Barreno_TS" not in df_detalle.columns:
        return None

    mask = pd.Series(True, index=df_detalle.index)
    if "Fuente" in df_detalle.columns:
        mask &= df_detalle["Fuente"].astype(str).str.upper().eq("ZDA")
    if "Jumbo" in df_detalle.columns:
        mask &= df_detalle["Jumbo"].astype(str).eq(str(jumbo))
    if "Ciclo" in df_detalle.columns:
        mask &= df_detalle["Ciclo"].astype(str).eq(str(ciclo))

    ts = pd.to_numeric(
        df_detalle.loc[mask, "Inicio_Barreno_TS"], errors="coerce"
    ).dropna()
    if ts.empty:
        return None

    d = datetime.fromtimestamp(int(ts.min()), tz=timezone.utc)
    return (d.hour * 3600 + d.minute * 60 + d.second) / 86400.0


def construir_bd_perfo(df_reportes, df_detalle):
    if df_reportes.empty or "Fuente" not in df_reportes.columns:
        return pd.DataFrame(columns=BD_PERFO_COLUMNS)

    op_map = _bd_operator_map(df_reportes)
    zda = df_reportes[
        df_reportes["Fuente"].astype(str).str.upper().eq("ZDA")
    ].copy()
    rows = []

    for _, r in zda.iterrows():
        dt = _bd_parse_datetime(r)
        seccion, roca, tipo_exp = _bd_plan_fields(r.get("Plan_Perforacion"))
        nivel, block, labor = _bd_labor_fields(r)

        auto = _bd_num(r.get("Auto_Total_Brazos_min"))
        manual = _bd_num(r.get("Manual_Total_Brazos_min"))

        if auto is None:
            a1 = _bd_num(r.get("Auto_Brazo1_min"))
            a2 = _bd_num(r.get("Auto_Brazo2_min"))
            if a1 is not None and a2 is not None:
                auto = a1 + a2

        if manual is None:
            m1 = _bd_num(r.get("Manual_Brazo1_min"))
            m2 = _bd_num(r.get("Manual_Brazo2_min"))
            if m1 is not None and m2 is not None:
                manual = m1 + m2

        movimiento = (
            auto + manual
            if auto is not None and manual is not None
            else None
        )
        taladros = _bd_num(r.get("Barrenos_Realizados"))
        t_automatico = auto * 60 / 26.5 if auto is not None else None
        t_manual = (
            taladros - t_automatico
            if taladros is not None and t_automatico is not None
            else None
        )
        obj_automatico = (
            t_automatico / taladros
            if taladros not in (None, 0) and t_automatico is not None
            else None
        )

        auto_b1 = _bd_num(r.get("Auto_Brazo1_min"))
        man_b1 = _bd_num(r.get("Manual_Brazo1_min"))
        auto_b2 = _bd_num(r.get("Auto_Brazo2_min"))
        man_b2 = _bd_num(r.get("Manual_Brazo2_min"))

        pct_auto_drill = (
            auto / (auto + manual)
            if auto is not None and manual is not None and (auto + manual) > 0
            else None
        )
        pct_auto_b1 = (
            auto_b1 / (auto_b1 + man_b1)
            if auto_b1 is not None and man_b1 is not None and (auto_b1 + man_b1) > 0
            else None
        )
        pct_auto_b2 = (
            auto_b2 / (auto_b2 + man_b2)
            if auto_b2 is not None and man_b2 is not None and (auto_b2 + man_b2) > 0
            else None
        )

        med_cut, prom_cut = _bd_cut_stats(
            df_detalle, r.get("Jumbo"), r.get("Ciclo")
        )
        primer_martillo = _bd_primer_martillo(
            df_detalle, r.get("Jumbo"), r.get("Ciclo")
        )

        conlabor = (
            f"{labor} {block} {nivel}"
            if labor and block is not None and nivel is not None
            else None
        )

        tiempo_perf_s = _bd_num(r.get("Tiempo_Perforacion_s"))

        rows.append({
            "MES": _bd_mes_operativo(dt),
            "SEMANA": None,
            "AÑO": dt.year if dt else None,
            "FECHA": datetime(dt.year, dt.month, dt.day) if dt else None,
            "TURNO": _bd_turno(dt),
            "JEFE DE TURNO": None,
            "OPERADOR": _bd_operador_exportado(r, op_map),
            "JUMBO": BD_JUMBO_ALIAS.get(r.get("Jumbo"), r.get("Jumbo")),
            "NIVEL": nivel,
            "BLOCK": block,
            "LABOR": labor,
            "TIPO DISPARO": r.get("Tipo_Disparo"),
            "SECCIÓN": seccion,
            "RMR": None,
            "TIP. DE ROCA": roca,
            "TIPO EXP.": tipo_exp,
            "Hora 1° taladro": _bd_hora_excel(dt),
            "Tiempo de perforación efec.": None,
            "Tiempo de perforación": (
                tiempo_perf_s / 86400.0 if tiempo_perf_s is not None else None
            ),
            "Tiempo de movimiento ": (
                movimiento / 1440.0 if movimiento is not None else None
            ),
            "T. mov. Manual": (
                manual / 1440.0 if manual is not None else None
            ),
            "T.mov. Autom.": (
                auto / 1440.0 if auto is not None else None
            ),
            "T. AUTOMÁTICO": t_automatico,
            "# TALADROS": taladros,
            "T. MANUAL": t_manual,
            "Av. Perf": None,
            "Av.Top": None,
            "Av.Real": None,
            "Eficiencia": None,
            "Obj. Automático": obj_automatico,
            "CONLABOR": conlabor,
            "Tipo Mat": None,
            "Ciclo": r.get("Ciclo"),
            "%AutoDrill": pct_auto_drill,
            "% Auto B1": pct_auto_b1,
            "% Auto B2": pct_auto_b2,
            "Mediana Cut (m)": med_cut,
            "Promedio Cut (m)": prom_cut,
            "Hora 1° martillo": primer_martillo,
        })

    return pd.DataFrame(rows, columns=BD_PERFO_COLUMNS)


@st.cache_data(show_spinner=False, max_entries=1)
def crear_excel_publicacion(df_bd_perfo, df_reportes, df_resumen):
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_bd_perfo.to_excel(writer, sheet_name="BD-PERFO", index=False)
        df_reportes.to_excel(writer, sheet_name="Resumen_Reportes", index=False)
        df_resumen.to_excel(writer, sheet_name="Resumen_Ciclos", index=False)

    buffer.seek(0)
    wb = load_workbook(buffer)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)

    for ws_any in wb.worksheets:
        ws_any.freeze_panes = "A2"
        if ws_any.max_row >= 1 and ws_any.max_column >= 1:
            ws_any.auto_filter.ref = ws_any.dimensions
            for c in ws_any[1]:
                c.fill = fill
                c.font = font
                c.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

    ws = wb["BD-PERFO"]

    formatos = {
        "D": "dd/mm/yyyy",
        "Q": "hh:mm:ss",
        "R": "[h]:mm",
        "S": "[h]:mm",
        "T": "[h]:mm",
        "U": "[h]:mm",
        "V": "[h]:mm",
        "W": "0",
        "Y": "0",
        "AC": "0.0%",
        "AD": "0.0%",
        "AH": "0.0%",
        "AI": "0.0%",
        "AJ": "0.0%",
        "AK": "0.00",
        "AL": "0.00",
        "AM": "hh:mm:ss",
    }

    for row in range(2, ws.max_row + 1):
        ws[f"C{row}"] = f"=YEAR(D{row})"
        ws[f"V{row}"] = f"=T{row}-U{row}"
        ws[f"W{row}"] = f"=V{row}*86400/26.5"
        ws[f"Y{row}"] = f"=X{row}-W{row}"
        ws[f"AC{row}"] = f"=AB{row}/Z{row}"
        ws[f"AD{row}"] = f"=W{row}/X{row}"
        ws[f"AE{row}"] = f'=K{row}&" "&J{row}&" "&I{row}'
        ws[f"AH{row}"] = f"=V{row}/T{row}"

        for col, fmt_code in formatos.items():
            ws[f"{col}{row}"].number_format = fmt_code

    widths = [
        12,10,8,12,10,18,18,10,10,10,16,24,12,10,14,16,16,24,22,22,
        18,18,16,12,14,12,12,12,12,18,24,14,10,14,14,14,16,16,16,
    ]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[
            ws.cell(row=1, column=idx).column_letter
        ].width = width

    for nombre_hoja in ["Resumen_Reportes", "Resumen_Ciclos"]:
        ws_tabla = wb[nombre_hoja]
        for col in ws_tabla.columns:
            letra = col[0].column_letter
            ancho = max(
                (len(str(c.value)) for c in col if c.value is not None),
                default=8,
            )
            ws_tabla.column_dimensions[letra].width = min(ancho + 2, 35)

    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida.getvalue()

# ==========================================================
# GRÁFICOS CONSOLIDADOS
# ==========================================================


def grafico_auto(df_auto: pd.DataFrame, mostrar_etiquetas: bool, mostrar_linea: bool):
    if df_auto.empty:
        return None
    df = df_auto.copy()
    # La población ya llega filtrada desde el bloque de automatización.
    # No se fuerza FRENTE aquí para que SELLADA y ESTOCADA Y/O CORRECCIONES
    # tengan efecto real en el filtro.
    df = df[df["Pct_Movimiento_Automatico_Brazos"].notna()].copy()
    df = asegurar_fechahora(df)
    if df.empty:
        return None

    fig = go.Figure()
    annotations = []
    points = []
    jumbos = list(df["Jumbo"].dropna().astype(str).unique())
    for idx, jumbo in enumerate(jumbos):
        g = df[df["Jumbo"].astype(str) == jumbo].sort_values("FechaHora")
        fig.add_trace(go.Scatter(
            x=g["FechaHora"], y=g["Pct_Movimiento_Automatico_Brazos"],
            mode="lines+markers" if mostrar_linea else "markers",
            name=jumbo,
            line=dict(
                width=3,
                color=COLORES[idx % len(COLORES)],
                shape="spline",
                smoothing=1.0,
            ) if mostrar_linea else None,
            marker=dict(
                size=9,
                color=COLORES[idx % len(COLORES)],
                line=dict(color="#ffffff", width=1.2),
            ),
            customdata=g[["Barrenos_Realizados", "Tipo_Roca"]].to_numpy(),
            hovertemplate=(f"{jumbo}<br>%{{x|%d/%m %H:%M}}<br>Automático: %{{y:.1f}}%"
                           "<br>Barrenos: %{customdata[0]} tal."
                           "<br>Tipo de roca: %{customdata[1]}<extra></extra>"),
        ))
        auto = pd.to_numeric(g["Auto_Total_Brazos_min"], errors="coerce").sum(min_count=1)
        manual = pd.to_numeric(g["Manual_Total_Brazos_min"], errors="coerce").sum(min_count=1)
        kpi = auto / (auto + manual) * 100 if pd.notna(auto) and pd.notna(manual) and (auto + manual) > 0 else None
        annotations.append(dict(
            xref="paper", yref="paper", x=0.01, y=1.18 - idx*0.08, xanchor="left",
            text=f"<b>{jumbo}</b> · Automático global: <b>{fmt(kpi,1,'%')}</b>",
            showarrow=False, bgcolor="rgba(255,255,255,0.92)", bordercolor="#dbe3ea", borderpad=5,
        ))
        if mostrar_etiquetas:
            for i, (_, r) in enumerate(g.iterrows()):
                points.append(dict(
                    x=r["FechaHora"], y=r["Pct_Movimiento_Automatico_Brazos"],
                    text=f"{r['Pct_Movimiento_Automatico_Brazos']:.1f}%<br>{int(r['Barrenos_Realizados']) if pd.notna(r.get('Barrenos_Realizados')) else '-'} tal.",
                    rank=i+idx,
                ))
    if mostrar_etiquetas:
        annotations.extend(smart_annotations(points, x_window_hours=18, y_window=4, font_size=11))
    fig.update_layout(**base_layout(
        470, annotations=annotations, margin=dict(l=78,r=35,t=120,b=72),
        yaxis=dict(title="Movimiento automático (%)", rangemode="tozero", gridcolor="#eef2f7"),
        xaxis=dict(title="Fecha", tickformat="%d/%m", gridcolor="#eef2f7"),
    ))
    return fig



def grafico_auto_por_operador(
    df_auto: pd.DataFrame,
    mostrar_etiquetas: bool,
    mostrar_linea: bool,
):
    """
    Evolución del movimiento automático por operador.

    - Cada línea/color representa un operador.
    - Cada punto representa un ciclo/round.
    - El % global mostrado en la leyenda se calcula a partir de la suma
      de tiempos automáticos y manuales del operador:
          Auto / (Auto + Manual)
      y no como promedio simple de porcentajes por ciclo.
    """
    if df_auto.empty:
        return None

    requeridas = {
        "Operador_Filtro",
        "Pct_Movimiento_Automatico_Brazos",
    }
    if not requeridas.issubset(df_auto.columns):
        return None

    df = df_auto.copy()

    df["Operador_Filtro"] = (
        df["Operador_Filtro"]
        .fillna("SIN DATO")
        .astype(str)
        .str.strip()
    )

    # Para una comparación realmente "por operador", no dibujar
    # registros donde no se logró identificar a la persona.
    df = df[
        df["Operador_Filtro"].ne("")
        & df["Operador_Filtro"].str.upper().ne("SIN DATO")
        & df["Pct_Movimiento_Automatico_Brazos"].notna()
    ].copy()

    df = asegurar_fechahora(df)
    if df.empty:
        return None

    operadores = sorted(
        df["Operador_Filtro"]
        .dropna()
        .astype(str)
        .unique()
    )

    fig = go.Figure()
    annotations = []
    points = []
    etiquetas_finales = []

    # Símbolo permite reconocer el jumbo sin competir con el color,
    # que queda reservado para distinguir operadores.
    jumbo_symbols = {
        "JUMB001": "circle",
        "JUMB002": "square",
    }

    # Paleta fija para facilitar el seguimiento visual de cada operador.
    # Cuchula se muestra en negro para diferenciarlo claramente de Osorio.
    color_map_operador = {
        "John Osorio": "#4F67F2",
        "Josue Rivera": "#F05A3A",
        "Nilton Celis": "#16C48A",
        "Rogelio Cuchula": "#111111",
    }

    for idx, operador in enumerate(operadores):
        color_operador = color_map_operador.get(
            str(operador),
            COLORES[idx % len(COLORES)],
        )

        g = df[
            df["Operador_Filtro"].astype(str).eq(str(operador))
        ].sort_values("FechaHora").copy()

        if g.empty:
            continue

        auto = pd.to_numeric(
            g.get("Auto_Total_Brazos_min"),
            errors="coerce",
        ).sum(min_count=1)

        manual = pd.to_numeric(
            g.get("Manual_Total_Brazos_min"),
            errors="coerce",
        ).sum(min_count=1)

        global_pct = (
            auto / (auto + manual) * 100
            if pd.notna(auto)
            and pd.notna(manual)
            and (auto + manual) > 0
            else None
        )

        ciclos = (
            g["Ciclo"]
            if "Ciclo" in g.columns
            else pd.Series(["-"] * len(g), index=g.index)
        )
        jumbos = (
            g["Jumbo"].fillna("-").astype(str)
            if "Jumbo" in g.columns
            else pd.Series(["-"] * len(g), index=g.index)
        )
        tipos = (
            g["Tipo_Disparo"].fillna("-").astype(str)
            if "Tipo_Disparo" in g.columns
            else pd.Series(["-"] * len(g), index=g.index)
        )
        rocas = (
            g["Tipo_Roca"].fillna("SIN DATO").astype(str)
            if "Tipo_Roca" in g.columns
            else pd.Series(["SIN DATO"] * len(g), index=g.index)
        )
        barrenos = (
            g["Barrenos_Realizados"]
            if "Barrenos_Realizados" in g.columns
            else pd.Series([None] * len(g), index=g.index)
        )

        custom = np.column_stack([
            ciclos.astype(object),
            jumbos.astype(object),
            tipos.astype(object),
            rocas.astype(object),
            barrenos.astype(object),
        ])

        symbols = [
            jumbo_symbols.get(str(j), "diamond")
            for j in jumbos
        ]

        nombre_leyenda = (
            f"{operador} · Global {global_pct:.1f}%"
            if global_pct is not None and pd.notna(global_pct)
            else operador
        )

        fig.add_trace(
            go.Scatter(
                x=g["FechaHora"],
                y=g["Pct_Movimiento_Automatico_Brazos"],
                mode="lines+markers" if mostrar_linea else "markers",
                name=nombre_leyenda,
                line=(
                    dict(
                        width=2.8,
                        color=color_operador,
                        shape="spline",
                        smoothing=1.0,
                    )
                    if mostrar_linea
                    else None
                ),
                marker=dict(
                    size=9,
                    symbol=symbols,
                    color=color_operador,
                    line=dict(
                        color="#ffffff",
                        width=1.1,
                    ),
                ),
                customdata=custom,
                hovertemplate=(
                    f"<b>{operador}</b>"
                    "<br>Fecha: %{x|%d/%m/%Y %H:%M}"
                    "<br>Jumbo: %{customdata[1]}"
                    "<br>Ciclo: %{customdata[0]}"
                    "<br>Automático: %{y:.1f}%"
                    "<br>Barrenos: %{customdata[4]}"
                    "<br>Tipo de disparo: %{customdata[2]}"
                    "<br>Tipo de roca: %{customdata[3]}"
                    "<extra></extra>"
                ),
            )
        )

        if mostrar_etiquetas:
            for i, (_, r) in enumerate(g.iterrows()):
                points.append({
                    "x": r["FechaHora"],
                    "y": r["Pct_Movimiento_Automatico_Brazos"],
                    "text": (
                        f"{operador.split()[-1]}<br>"
                        f"{r['Pct_Movimiento_Automatico_Brazos']:.1f}%"
                    ),
                    "rank": i + idx * 100,
                })

        # Guardar la etiqueta final para posicionarla después.
        # El ajuste conjunto permite evitar superposición entre apellidos.
        ultimo = g.iloc[-1]
        etiquetas_finales.append({
            "x": ultimo["FechaHora"],
            "y": float(ultimo["Pct_Movimiento_Automatico_Brazos"]),
            "texto": operador.split()[-1],
            "color": color_operador,
        })

    # ------------------------------------------------------
    # Etiquetas finales sin superposición
    # ------------------------------------------------------
    # Si dos o más operadores terminan con valores muy próximos,
    # sus apellidos se desplazan verticalmente en píxeles.
    # Los puntos y las curvas permanecen exactamente en su valor real.
    if etiquetas_finales:
        etiquetas_ordenadas = sorted(
            etiquetas_finales,
            key=lambda e: e["y"],
            reverse=True,
        )

        # Umbral en puntos porcentuales para considerar que dos etiquetas
        # podrían superponerse visualmente.
        umbral_pp = 6.0
        clusters = []
        cluster_actual = []

        for etiqueta in etiquetas_ordenadas:
            if not cluster_actual:
                cluster_actual = [etiqueta]
                continue

            if abs(cluster_actual[-1]["y"] - etiqueta["y"]) <= umbral_pp:
                cluster_actual.append(etiqueta)
            else:
                clusters.append(cluster_actual)
                cluster_actual = [etiqueta]

        if cluster_actual:
            clusters.append(cluster_actual)

        # Desplazamientos simétricos alrededor de la posición real.
        # Ejemplos:
        # 2 etiquetas -> +11 / -11 px
        # 3 etiquetas -> +22 / 0 / -22 px
        # 4 etiquetas -> +33 / +11 / -11 / -33 px
        for cluster in clusters:
            n = len(cluster)
            paso_px = 22
            centro = (n - 1) / 2.0

            for pos, etiqueta in enumerate(cluster):
                yshift = int(round((centro - pos) * paso_px))

                annotations.append(
                    dict(
                        x=etiqueta["x"],
                        y=etiqueta["y"],
                        xref="x",
                        yref="y",
                        text=f'<b>{etiqueta["texto"]}</b>',
                        showarrow=False,
                        xanchor="left",
                        yanchor="middle",
                        xshift=14,
                        yshift=yshift,
                        font=dict(
                            size=11,
                            color=etiqueta["color"],
                        ),
                        bgcolor="rgba(255,255,255,0.90)",
                        bordercolor="rgba(0,0,0,0)",
                        borderpad=2,
                    )
                )

    if mostrar_etiquetas and points:
        annotations.extend(
            smart_annotations(
                points,
                x_window_hours=18,
                y_window=4,
                font_size=10,
            )
        )

    fig.update_layout(
        **base_layout(
            500,
            annotations=annotations,
            margin=dict(
                l=78,
                r=145,
                t=55,
                b=82,
            ),
            yaxis=dict(
                title="Movimiento automático (%)",
                rangemode="tozero",
                gridcolor="#eef2f7",
            ),
            xaxis=dict(
                title="Fecha",
                tickformat="%d/%m",
                gridcolor="#eef2f7",
            ),
            legend=dict(
                orientation="h",
                y=-0.20,
                x=0,
            ),
            hovermode="closest",
        )
    )

    return fig


def grafico_brazos(df_auto: pd.DataFrame, jumbo: str, mostrar_etiquetas: bool):
    df = df_auto[df_auto["Jumbo"].astype(str) == str(jumbo)].copy()
    # La población ya llega filtrada desde el bloque de automatización.
    df = df[df["Pct_Automatico_Brazo1"].notna() | df["Pct_Automatico_Brazo2"].notna()].copy()
    df = asegurar_fechahora(df)
    if df.empty:
        return None

    fig = go.Figure()
    annotations = []
    points = []
    series = [
        ("Brazo 1", "Pct_Automatico_Brazo1", "Auto_Brazo1_min", "Manual_Brazo1_min", 0),
        ("Brazo 2", "Pct_Automatico_Brazo2", "Auto_Brazo2_min", "Manual_Brazo2_min", 1),
    ]
    for nombre, col, auto_col, man_col, idx in series:
        g = df[df[col].notna()].sort_values("FechaHora")
        if g.empty:
            continue
        auto = pd.to_numeric(g[auto_col], errors="coerce").sum(min_count=1)
        manual = pd.to_numeric(g[man_col], errors="coerce").sum(min_count=1)
        global_pct = auto / (auto + manual) * 100 if pd.notna(auto) and pd.notna(manual) and (auto + manual) > 0 else None
        fig.add_trace(go.Scatter(
            x=g["FechaHora"], y=g[col], mode="lines+markers",
            name=f"{nombre} · Global {fmt(global_pct,1,'%')}",
            line=dict(width=2.5, color=COLORES[idx], shape="spline"), marker=dict(size=8),
            hovertemplate=f"{nombre}<br>%{{x|%d/%m %H:%M}}<br>Automático: %{{y:.1f}}%<extra></extra>",
        ))
        annotations.append(dict(
            xref="paper", yref="paper", x=0.01 + idx*0.25, y=1.16, xanchor="left",
            text=f"<b>{nombre} global: {fmt(global_pct,1,'%')}</b>", showarrow=False,
            bgcolor="rgba(255,255,255,0.92)", bordercolor="#dbe3ea", borderpad=5,
        ))
        if mostrar_etiquetas:
            for i, (_, r) in enumerate(g.iterrows()):
                points.append(dict(x=r["FechaHora"], y=r[col], text=f"{r[col]:.1f}%", rank=i+idx*100))
    if mostrar_etiquetas:
        annotations.extend(smart_annotations(points, x_window_hours=18, y_window=4, font_size=10))
    fig.update_layout(**base_layout(
        430, title=dict(text=f"<b>{jumbo}</b>", x=0.01), annotations=annotations,
        margin=dict(l=72,r=30,t=92,b=72),
        yaxis=dict(title="Movimiento automático (%)", rangemode="tozero", gridcolor="#eef2f7"),
        xaxis=dict(title="Fecha", tickformat="%d/%m", gridcolor="#eef2f7"),
    ))
    return fig


def preparar_cut(df_resumen: pd.DataFrame, df_reportes: pd.DataFrame) -> pd.DataFrame:
    if df_resumen.empty:
        return pd.DataFrame()
    cut = df_resumen[df_resumen["Tipo"].astype(str).str.lower() == "cut"].copy()
    cut = cut[cut["Mediana"].notna()].copy()
    if cut.empty:
        return cut
    cols = [c for c in ["Jumbo","Ciclo","Fecha_Inicio","Tipo_Disparo","Tipo_Roca","Operador_Filtro"] if c in df_reportes.columns]
    tipos = df_reportes[cols].drop_duplicates(subset=[c for c in ["Jumbo","Ciclo","Fecha_Inicio"] if c in cols])
    cut = cut.merge(tipos, on=["Jumbo","Ciclo","Fecha_Inicio"], how="left")
    cut["Tipo_Disparo"] = cut["Tipo_Disparo"].fillna("SIN CLASIFICAR")
    return asegurar_fechahora(cut)



def grafico_cut(
    df_cut: pd.DataFrame,
    jumbos_visibles,
    tipos_visibles,
    rocas_visibles,
    operadores_visibles,
    mostrar_etiquetas: bool,
):
    if df_cut.empty:
        return None

    df = df_cut[
        df_cut["Jumbo"].astype(str).isin([str(x) for x in jumbos_visibles])
        & df_cut["Tipo_Disparo"].isin(tipos_visibles)
        & df_cut["Tipo_Roca"].isin(rocas_visibles)
        & df_cut["Operador_Filtro"].isin(operadores_visibles)
    ].copy()

    if df.empty:
        return None

    all_jumbos = sorted(df_cut["Jumbo"].dropna().astype(str).unique())
    visibles = sorted(df["Jumbo"].dropna().astype(str).unique())

    fig = go.Figure()
    points = []
    annotations = []

    for pos_visible, jumbo in enumerate(visibles):
        idx = all_jumbos.index(jumbo)
        g = df[df["Jumbo"].astype(str) == jumbo].sort_values("FechaHora")
        custom = g[["Ciclo", "Tipo_Disparo"]].to_numpy()

        fig.add_trace(go.Scatter(
            x=g["FechaHora"],
            y=g["Mediana"],
            mode="lines+markers",
            name=jumbo,
            customdata=custom,
            line=dict(
                width=3,
                color=COLORES[idx % len(COLORES)],
                shape="spline",
            ),
            marker=dict(size=8),
            hovertemplate=(
                f"{jumbo}<br>%{{x|%d/%m %H:%M}}"
                "<br>Ciclo: %{customdata[0]}"
                "<br>Tipo: %{customdata[1]}"
                "<br>Mediana Cut: %{y:.2f} m<extra></extra>"
            ),
        ))

        # Global coherente con la gráfica:
        # mediana de las medianas Cut de los ciclos visibles.
        global_cut = pd.to_numeric(g["Mediana"], errors="coerce").dropna().median()
        if pd.notna(global_cut):
            annotations.append(dict(
                xref="paper",
                yref="paper",
                x=0.01 + pos_visible * 0.32,
                y=1.16,
                xanchor="left",
                showarrow=False,
                text=f"<b>{jumbo}</b> · Longitud global Cut: <b>{global_cut:.2f} m</b>",
                font=dict(
                    size=12,
                    color=COLORES[idx % len(COLORES)],
                ),
                bgcolor="rgba(255,255,255,0.95)",
                bordercolor="#dbe3ea",
                borderpad=5,
            ))

        if mostrar_etiquetas:
            for i, (_, r) in enumerate(g.iterrows()):
                points.append(dict(
                    x=r["FechaHora"],
                    y=r["Mediana"],
                    text=f"{r['Mediana']:.2f} m",
                    rank=i + idx * 50,
                ))

    if mostrar_etiquetas:
        annotations.extend(
            smart_annotations(
                points,
                x_window_hours=18,
                y_window=0.18,
                font_size=11,
            )
        )

    fig.update_layout(**base_layout(
        450,
        annotations=annotations,
        margin=dict(l=80, r=30, t=88, b=70),
        yaxis=dict(
            title="Mediana de longitud perforada Cut (m)",
            gridcolor="#eef2f7",
        ),
        xaxis=dict(
            title="Fecha",
            tickformat="%d/%m",
            gridcolor="#eef2f7",
        ),
    ))
    return fig

# ==========================================================
# ZDA: RESÚMENES Y TIMELINE
# ==========================================================


def _utc_dt(ts):
    return datetime.fromtimestamp(int(ts), tz=timezone.utc)


def zda_operational_date(ts):
    d = _utc_dt(ts)
    base = datetime(d.year, d.month, d.day, tzinfo=timezone.utc)
    if d.hour < 7:
        base -= timedelta(days=1)
    return base


def zda_operational_hour(ts, op_date):
    d = _utc_dt(ts)
    day_diff = (datetime(d.year,d.month,d.day,tzinfo=timezone.utc) - op_date).days
    return d.hour + d.minute/60 + d.second/3600 + 24*day_diff


def zda_turno(ts):
    d = _utc_dt(ts)
    h = d.hour + d.minute/60 + d.second/3600
    if 7 <= h < 19:
        return "Día", h
    return "Noche", h+24 if h < 7 else h


def fmt_hora_decimal(h):
    if h is None or pd.isna(h):
        return "-"
    total = int(round(float(h)*60)) % (24*60)
    return f"{total//60:02d}:{total%60:02d}"


def resumen_turnos_zda(rows: pd.DataFrame) -> pd.DataFrame:
    if rows.empty:
        return pd.DataFrame()
    work = rows.copy()
    work["_turno"] = work["Inicio_Perforacion_TS"].apply(lambda x: zda_turno(x)[0])
    work["_opHour"] = work["Inicio_Perforacion_TS"].apply(lambda x: zda_turno(x)[1])
    work["_opDate"] = work["Inicio_Perforacion_TS"].apply(zda_operational_date)

    # Primer round por jumbo, fecha operativa y turno para indicadores de inicio.
    first_idx = work.groupby(["Jumbo","_opDate","_turno"])["Inicio_Perforacion_TS"].idxmin()
    first = work.loc[first_idx].copy()
    salida = []
    for jumbo in sorted(work["Jumbo"].dropna().astype(str).unique()):
        allj = work[work["Jumbo"].astype(str)==jumbo]
        firstj = first[first["Jumbo"].astype(str)==jumbo]
        day_all = allj[allj["_turno"]=="Día"]
        night_all = allj[allj["_turno"]=="Noche"]
        day = firstj[firstj["_turno"]=="Día"]
        night = firstj[firstj["_turno"]=="Noche"]
        salida.append({
            "Jumbo": jumbo,
            "Ciclos día": len(day_all),
            "Inicio prom. día": fmt_hora_decimal(day["_opHour"].mean()) if not day.empty else "-",
            "Inicio más temprano día": fmt_hora_decimal(day["_opHour"].min()) if not day.empty else "-",
            "Inicio más tarde día": fmt_hora_decimal(day["_opHour"].max()) if not day.empty else "-",
            "Ciclos noche": len(night_all),
            "Inicio prom. noche": fmt_hora_decimal(night["_opHour"].mean()) if not night.empty else "-",
            "Inicio más temprano noche": fmt_hora_decimal(night["_opHour"].min()) if not night.empty else "-",
            "Inicio más tarde noche": fmt_hora_decimal(night["_opHour"].max()) if not night.empty else "-",
        })
    return pd.DataFrame(salida)


def resumen_fin_turnos_zda(rows: pd.DataFrame) -> pd.DataFrame:
    """
    Resume la hora de término del último round por jumbo y turno.

    - Ciclos día/noche: cuenta todos los rounds iniciados en cada turno.
    - Fin prom./más temprano/más tarde: usa únicamente el FIN DEL ÚLTIMO
      round de cada Jumbo + Fecha operativa + Turno.

    Para el turno noche, la hora se calcula sobre una escala continua
    19:00 -> 07:00 (por ejemplo, 02:00 = 26.0) y luego se vuelve a mostrar
    como hora reloj mediante fmt_hora_decimal().
    """
    if rows.empty:
        return pd.DataFrame()

    requeridas = {
        "Jumbo",
        "Inicio_Perforacion_TS",
        "Fin_Perforacion_TS",
    }
    if not requeridas.issubset(rows.columns):
        return pd.DataFrame()

    work = rows[
        rows["Inicio_Perforacion_TS"].notna()
        & rows["Fin_Perforacion_TS"].notna()
        & rows["Jumbo"].notna()
    ].copy()

    if work.empty:
        return pd.DataFrame()

    # El turno y la fecha operativa se determinan por el INICIO del round.
    work["_turno"] = work["Inicio_Perforacion_TS"].apply(
        lambda x: zda_turno(x)[0]
    )
    work["_opDate"] = work["Inicio_Perforacion_TS"].apply(
        zda_operational_date
    )

    # Hora de fin continua respecto de la fecha operativa.
    # Esto evita errores al cruzar medianoche en el turno noche.
    work["_finHour"] = work.apply(
        lambda r: zda_operational_hour(
            r["Fin_Perforacion_TS"],
            r["_opDate"],
        ),
        axis=1,
    )

    # Último round por Jumbo + Fecha operativa + Turno.
    last_idx = work.groupby(
        ["Jumbo", "_opDate", "_turno"]
    )["Fin_Perforacion_TS"].idxmax()
    last = work.loc[last_idx].copy()

    salida = []

    for jumbo in sorted(work["Jumbo"].dropna().astype(str).unique()):
        allj = work[work["Jumbo"].astype(str) == jumbo]
        lastj = last[last["Jumbo"].astype(str) == jumbo]

        day_all = allj[allj["_turno"] == "Día"]
        night_all = allj[allj["_turno"] == "Noche"]

        day = lastj[lastj["_turno"] == "Día"]
        night = lastj[lastj["_turno"] == "Noche"]

        salida.append({
            "Jumbo": jumbo,
            "Ciclos día": len(day_all),
            "Fin prom. día": (
                fmt_hora_decimal(day["_finHour"].mean())
                if not day.empty else "-"
            ),
            "Fin más temprano día": (
                fmt_hora_decimal(day["_finHour"].min())
                if not day.empty else "-"
            ),
            "Fin más tarde día": (
                fmt_hora_decimal(day["_finHour"].max())
                if not day.empty else "-"
            ),
            "Ciclos noche": len(night_all),
            "Fin prom. noche": (
                fmt_hora_decimal(night["_finHour"].mean())
                if not night.empty else "-"
            ),
            "Fin más temprano noche": (
                fmt_hora_decimal(night["_finHour"].min())
                if not night.empty else "-"
            ),
            "Fin más tarde noche": (
                fmt_hora_decimal(night["_finHour"].max())
                if not night.empty else "-"
            ),
        })

    return pd.DataFrame(salida)


def resumen_tipos_zda(rows: pd.DataFrame) -> pd.DataFrame:
    if rows.empty:
        return pd.DataFrame()
    jumbos = sorted(rows["Jumbo"].dropna().astype(str).unique())
    salida = []
    for tipo in TIPOS_DISPARO:
        r = {"Tipo": tipo, "Total": int((rows["Tipo_Disparo"]==tipo).sum())}
        for j in jumbos:
            r[j] = int(((rows["Tipo_Disparo"]==tipo) & (rows["Jumbo"].astype(str)==j)).sum())
        if r["Total"] > 0:
            salida.append(r)
    return pd.DataFrame(salida)


def grafico_zda_timeline(rows: pd.DataFrame, mostrar_etiquetas: bool):
    if rows.empty:
        return None
    rows = rows.sort_values("Inicio_Perforacion_TS").copy()
    all_jumbos = sorted(rows["Jumbo"].dropna().astype(str).unique())
    fig = go.Figure()
    annotations = []
    legend_done = set()

    rows["_opDate"] = rows["Inicio_Perforacion_TS"].apply(zda_operational_date)
    for op_date, grupo in rows.groupby("_opDate", sort=True):
        grupo = grupo.sort_values("Inicio_Perforacion_TS")
        n = len(grupo)
        for i, (_, r) in enumerate(grupo.iterrows()):
            offset = (i - (n-1)/2) * 0.11
            x = op_date + timedelta(days=offset)
            y1 = zda_operational_hour(r["Inicio_Perforacion_TS"], op_date)
            y2 = zda_operational_hour(r["Fin_Perforacion_TS"], op_date)
            jumbo = str(r["Jumbo"])
            idx = all_jumbos.index(jumbo)
            n_b = r.get("Barrenos_ZDA") if pd.notna(r.get("Barrenos_ZDA")) else r.get("Barrenos_Realizados")
            custom = [[r.get("Ciclo"), n_b, r.get("Inicio_Perforacion"), r.get("Fin_Perforacion"), r.get("Tiempo_Perforacion_hms"), r.get("Labor") or "-", r.get("Tipo_Disparo")]] * 2
            fig.add_trace(go.Scatter(
                x=[x,x], y=[y1,y2], mode="lines+markers", name=jumbo,
                legendgroup=jumbo, showlegend=jumbo not in legend_done,
                line=dict(
                    width=12,
                    color=COLORES[idx%len(COLORES)],
                ),
                marker=dict(
                    size=[12, 7],
                    symbol=["diamond", "circle"],
                    color=COLORES[idx%len(COLORES)],
                    line=dict(color="#ffffff", width=1.4),
                ),
                customdata=custom,
                hovertemplate=(f"{jumbo} · Ciclo %{{customdata[0]}}<br>Tipo: %{{customdata[6]}}"
                               "<br>Inicio: %{customdata[2]}<br>Fin: %{customdata[3]}<br>Tiempo: %{customdata[4]}"
                               "<br>Barrenos: %{customdata[1]} B<br>Labor: %{customdata[5]}<extra></extra>"),
            ))
            legend_done.add(jumbo)
            if mostrar_etiquetas:
                annotations.append(dict(
                    x=x, y=(y1+y2)/2, xref="x", yref="y", showarrow=False, xshift=18,
                    text=f"C{r.get('Ciclo')} · {r.get('Tiempo_Perforacion_hms') or '-'} | {int(n_b) if pd.notna(n_b) else '-'}B",
                    font=dict(size=10,color="#334155"), bgcolor="rgba(255,255,255,.9)",
                    bordercolor="#dbe3ea", borderpad=3, xanchor="left",
                ))

    # Eje horario cada 1 hora: 07:00 -> 07:00 del día siguiente.
    tickvals = list(range(7, 32))
    ticktext = [f"{v%24:02d}:00" for v in tickvals]
    op_dates = sorted(rows["_opDate"].unique())

    # Jerarquía visual del eje horario:
    # - grilla base cada 1 hora (muy tenue)
    # - línea principal cada 2 horas
    # - 19:00 se reserva para el cambio de turno
    major_hour_lines = [
        dict(
            type="line",
            xref="paper",
            x0=0,
            x1=1,
            yref="y",
            y0=h,
            y1=h,
            line=dict(
                color="#cbd5e1",
                width=1.15,
            ),
            layer="below",
        )
        for h in range(7, 32, 2)
        if h != 19
    ]

    fig.update_layout(**base_layout(
        720, annotations=annotations, margin=dict(l=80,r=150,t=55,b=72),
        xaxis=dict(title="Fecha operativa", tickvals=op_dates,
                   ticktext=[pd.Timestamp(x).strftime("%d/%m") for x in op_dates], gridcolor="#eef2f7"),
        yaxis=dict(
            title="Hora",
            range=[31.2,6.8],
            tickmode="array",
            tickvals=tickvals,
            ticktext=ticktext,
            gridcolor="#edf2f7",
            gridwidth=0.55,
        ),
        shapes=[
            dict(
                type="rect",
                xref="paper",
                x0=0,
                x1=1,
                yref="y",
                y0=7,
                y1=19,
                fillcolor="rgba(37,99,235,.035)",
                line=dict(width=0),
                layer="below",
            ),
            dict(
                type="rect",
                xref="paper",
                x0=0,
                x1=1,
                yref="y",
                y0=19,
                y1=31,
                fillcolor="rgba(15,23,42,.035)",
                line=dict(width=0),
                layer="below",
            ),
            *major_hour_lines,
            dict(
                type="line",
                xref="paper",
                x0=0,
                x1=1,
                yref="y",
                y0=19,
                y1=19,
                line=dict(
                    color="#7c8da6",
                    width=1.3,
                    dash="dot",
                ),
            ),
        ],
    ))
    return fig


def _turno_inicio_ciclo(ts):
    d = _utc_dt(ts)
    h = d.hour + d.minute / 60 + d.second / 3600
    return "Día" if 7 <= h < 19 else "Noche"


def preparar_timeline_ciclos_turno(rows: pd.DataFrame) -> pd.DataFrame:
    """
    Prepara ciclos físicos únicos para el timeline por día/turno.

    Identidad del ciclo:
        Jumbo + número de Ciclo

    Cada fila resultante representa UN round/frente perforado y contiene:
        - fecha operativa,
        - turno de inicio,
        - hora relativa de inicio,
        - hora relativa de fin,
        - duración,
        - jumbo,
        - número de ciclo.

    No se asigna 1.er/2.º/3.º ciclo. Si existen dos rounds reales del mismo
    jumbo dentro de un turno, aparecerán naturalmente como dos segmentos.
    """
    if rows.empty:
        return pd.DataFrame()

    requeridas = {
        "Inicio_Perforacion_TS",
        "Fin_Perforacion_TS",
        "Jumbo",
        "Ciclo",
    }
    if not requeridas.issubset(rows.columns):
        return pd.DataFrame()

    work = rows[
        rows["Inicio_Perforacion_TS"].notna()
        & rows["Fin_Perforacion_TS"].notna()
        & rows["Jumbo"].notna()
        & rows["Ciclo"].notna()
    ].copy()

    if work.empty:
        return pd.DataFrame()

    work["_Inicio"] = pd.to_numeric(
        work["Inicio_Perforacion_TS"],
        errors="coerce",
    )
    work["_Fin"] = pd.to_numeric(
        work["Fin_Perforacion_TS"],
        errors="coerce",
    )

    work = work[
        work["_Inicio"].notna()
        & work["_Fin"].notna()
        & (work["_Fin"] >= work["_Inicio"])
    ].copy()

    if work.empty:
        return pd.DataFrame()

    work["_Jumbo_Key"] = (
        work["Jumbo"]
        .astype(str)
        .str.strip()
    )
    work["_Ciclo_Key"] = (
        work["Ciclo"]
        .astype(str)
        .str.strip()
    )

    work["_Duracion_s"] = (
        work["_Fin"] - work["_Inicio"]
    )

    # Un mismo número de ciclo del mismo jumbo se considera el mismo round.
    # Si aparece duplicado, conservamos el registro con mayor ventana temporal.
    work = (
        work.sort_values(
            [
                "_Jumbo_Key",
                "_Ciclo_Key",
                "_Duracion_s",
            ],
            ascending=[True, True, False],
        )
        .drop_duplicates(
            subset=[
                "_Jumbo_Key",
                "_Ciclo_Key",
            ],
            keep="first",
        )
        .copy()
    )

    salida = []

    for _, r in work.iterrows():
        start_ts = float(r["_Inicio"])
        end_ts = float(r["_Fin"])

        op_date = zda_operational_date(start_ts)
        turno = _turno_inicio_ciclo(start_ts)

        if turno == "Día":
            shift_start = op_date + timedelta(hours=7)
        else:
            shift_start = op_date + timedelta(hours=19)

        shift_start_ts = shift_start.timestamp()

        x_inicio = (
            start_ts - shift_start_ts
        ) / 3600.0

        x_fin = (
            end_ts - shift_start_ts
        ) / 3600.0

        salida.append({
            "Fecha_Operativa_DT": op_date,
            "Fecha_Operativa": op_date.strftime("%d/%m/%Y"),
            "Turno": turno,
            "Dia_Turno": (
                f"{op_date.strftime('%d/%m')} · {turno}"
            ),
            "Jumbo": str(r.get("Jumbo") or "-"),
            "Ciclo": r.get("Ciclo"),
            "X_Inicio": x_inicio,
            "X_Fin": x_fin,
            "Inicio": (
                r.get("Inicio_Perforacion")
                or _utc_dt(start_ts).strftime("%d/%m/%Y %H:%M:%S")
            ),
            "Fin": (
                r.get("Fin_Perforacion")
                or _utc_dt(end_ts).strftime("%d/%m/%Y %H:%M:%S")
            ),
            "Duracion": (
                r.get("Tiempo_Perforacion_hms")
                or format_duration_hms(end_ts - start_ts)
            ),
            "Duracion_h": (
                end_ts - start_ts
            ) / 3600.0,
            "Tipo_Disparo": r.get("Tipo_Disparo") or "-",
            "Tipo_Roca": r.get("Tipo_Roca") or "SIN DATO",
            "Labor": r.get("Labor") or "-",
            "Operador": (
                r.get("Operador_ZDA")
                or r.get("Operador")
                or "-"
            ),
            "Barrenos": (
                r.get("Barrenos_ZDA")
                if pd.notna(r.get("Barrenos_ZDA"))
                else r.get("Barrenos_Realizados")
            ),
            "Sobrepasa_Turno": bool(x_fin > 12),
        })

    if not salida:
        return pd.DataFrame()

    return (
        pd.DataFrame(salida)
        .sort_values(
            [
                "Fecha_Operativa_DT",
                "Turno",
                "Jumbo",
                "X_Inicio",
            ]
        )
        .reset_index(drop=True)
    )


def grafico_timeline_ciclos_turno(
    ciclos: pd.DataFrame,
    turno: str,
    solo_puntos_inicio: bool = False,
    solo_primer_inicio: bool = False,
):
    """
    Timeline horizontal por fecha operativa.

    X:
        horas transcurridas desde el inicio del turno.

    Y:
        fecha operativa + turno.

    Dentro de cada fila:
        JUMB001 se dibuja ligeramente arriba,
        JUMB002 ligeramente abajo.

    Cada round real es un segmento completo Inicio -> Fin.
    Si un jumbo ejecuta dos rounds, aparecen dos segmentos consecutivos
    sobre la misma pista de ese jumbo.

    Por defecto se muestran barras horizontales Inicio -> Fin.
    Si solo_puntos_inicio=True, las barras se ocultan completamente y
    se muestran únicamente los puntos de inicio para analizar patrones horarios.
    Si además solo_primer_inicio=True, se muestra únicamente el primer
    inicio de cada equipo por cada combinación fecha operativa + turno.
    """
    if ciclos.empty:
        return None

    g = ciclos[
        ciclos["Turno"].eq(turno)
    ].copy()

    if g.empty:
        return None

    # Vista de patrones de arranque:
    # cuando se pide solo el primer inicio, conservar únicamente
    # el round que empezó más temprano para cada Jumbo + fecha operativa + turno.
    if solo_puntos_inicio and solo_primer_inicio:
        g = (
            g.sort_values(
                [
                    "Fecha_Operativa_DT",
                    "Jumbo",
                    "X_Inicio",
                    "Ciclo",
                ]
            )
            .drop_duplicates(
                subset=[
                    "Fecha_Operativa_DT",
                    "Turno",
                    "Jumbo",
                ],
                keep="first",
            )
            .copy()
        )

    fechas = sorted(
        g["Fecha_Operativa_DT"].dropna().unique()
    )

    if not fechas:
        return None

    y_map = {
        pd.Timestamp(fecha): i
        for i, fecha in enumerate(fechas)
    }

    offsets = {
        "JUMB001": -0.16,
        "JUMB002": 0.16,
    }

    estilos = {
        "JUMB001": {
            "line_color": "#64748b",
            "dash": "solid",
            "marker_color": "#64748b",
            "legend_name": "JUMB001",
        },
        "JUMB002": {
            "line_color": "#111827",
            "dash": "solid",
            "marker_color": "#111827",
            "legend_name": "JUMB002",
        },
    }

    fig = go.Figure()
    legend_done = set()

    # Eje X fijo para mantener siempre la misma referencia visual.
    # 0 = inicio del turno y 12 = fin del turno.
    x_max = 12.0

    for _, r in g.iterrows():
        jumbo = str(r["Jumbo"])
        estilo = estilos.get(
            jumbo,
            estilos["JUMB001"],
        )

        fecha_ts = pd.Timestamp(
            r["Fecha_Operativa_DT"]
        )
        y_base = y_map[fecha_ts]
        y = y_base + offsets.get(jumbo, 0)

        x1 = float(r["X_Inicio"])
        x2 = float(r["X_Fin"])

        custom = [[
            jumbo,
            r.get("Ciclo"),
            r.get("Inicio"),
            r.get("Fin"),
            r.get("Duracion"),
            r.get("Tipo_Disparo"),
            r.get("Tipo_Roca"),
            r.get("Labor"),
            r.get("Barrenos"),
            r.get("Fecha_Operativa"),
            "Sí" if r.get("Sobrepasa_Turno") else "No",
            r.get("Operador") or "-",
        ]] * 2

        if solo_puntos_inicio:
            # Modo análisis de inicios:
            # ocultar completamente la barra y mostrar solo el inicio.
            if jumbo == "JUMB001":
                # En símbolos "open", Plotly usa marker.color como
                # color principal del contorno. No debe ser blanco.
                symbol = "square-open"
                marker_color = "#64748b"
                marker_line_color = "#64748b"
            else:
                symbol = "square"
                marker_color = "#111827"
                marker_line_color = "#111827"

            fig.add_trace(
                go.Scatter(
                    x=[x1],
                    y=[y],
                    mode="markers",
                    name=estilo["legend_name"],
                    legendgroup=jumbo,
                    showlegend=jumbo not in legend_done,
                    marker=dict(
                        size=11,
                        symbol=symbol,
                        color=marker_color,
                        line=dict(
                            color=marker_line_color,
                            width=1.8,
                        ),
                    ),
                    customdata=[custom[0]],
                    hovertemplate=(
                        "<b>%{customdata[0]}</b>"
                        "<br>Ciclo / round: %{customdata[1]}"
                        "<br>Fecha operativa: %{customdata[9]}"
                        "<br>Inicio: %{customdata[2]}"
                        "<br>Fin: %{customdata[3]}"
                        "<br>Duración: %{customdata[4]}"
                        "<br>Tipo: %{customdata[5]}"
                        "<br>Tipo de roca: %{customdata[6]}"
                        "<br>Labor: %{customdata[7]}"
                        "<br>Operador: %{customdata[11]}"
                        "<br>Barrenos: %{customdata[8]}"
                        "<br>Sobrepasa turno: %{customdata[10]}"
                        "<extra></extra>"
                    ),
                )
            )
        else:
            # Modo timeline completo, sincronizado con la vista de puntos:
            # JUMB001 = barra hueca con contorno cerrado.
            # JUMB002 = barra negra sólida.
            hover_barra = (
                "<b>%{customdata[0]}</b>"
                "<br>Ciclo / round: %{customdata[1]}"
                "<br>Fecha operativa: %{customdata[9]}"
                "<br>Inicio: %{customdata[2]}"
                "<br>Fin: %{customdata[3]}"
                "<br>Duración: %{customdata[4]}"
                "<br>Tipo: %{customdata[5]}"
                "<br>Tipo de roca: %{customdata[6]}"
                "<br>Labor: %{customdata[7]}"
                "<br>Operador: %{customdata[11]}"
                "<br>Barrenos: %{customdata[8]}"
                "<br>Sobrepasa turno: %{customdata[10]}"
                "<extra></extra>"
            )

            duracion = max(float(x2 - x1), 0.02)
            ancho_barra = 0.15
            show_legend_actual = jumbo not in legend_done

            if jumbo == "JUMB001":
                fig.add_trace(
                    go.Bar(
                        x=[duracion],
                        y=[y],
                        base=[x1],
                        orientation="h",
                        width=ancho_barra,
                        name=estilo["legend_name"],
                        legendgroup=jumbo,
                        showlegend=show_legend_actual,
                        marker=dict(
                            color="rgba(255,255,255,0)",
                            line=dict(
                                color="#64748b",
                                width=1.8,
                            ),
                        ),
                        customdata=[custom[0]],
                        hovertemplate=hover_barra,
                    )
                )
            else:
                fig.add_trace(
                    go.Bar(
                        x=[duracion],
                        y=[y],
                        base=[x1],
                        orientation="h",
                        width=ancho_barra,
                        name=estilo["legend_name"],
                        legendgroup=jumbo,
                        showlegend=show_legend_actual,
                        marker=dict(
                            color="#111827",
                            line=dict(
                                color="#111827",
                                width=0.8,
                            ),
                        ),
                        customdata=[custom[0]],
                        hovertemplate=hover_barra,
                    )
                )

        legend_done.add(jumbo)

    tickvals = list(range(0, 13))

    if turno == "Día":
        ticktext = [f"{(7 + h) % 24:02d}:00" for h in tickvals]
        titulo = "Turno Día · 07:00–19:00"
    else:
        ticktext = [f"{(19 + h) % 24:02d}:00" for h in tickvals]
        titulo = "Turno Noche · 19:00–07:00"

    y_tickvals = [y_map[pd.Timestamp(fecha)] for fecha in fechas]
    y_ticktext = [f"{pd.Timestamp(fecha).strftime('%d/%m')} · {turno}" for fecha in fechas]

    height = max(420, min(980, 150 + len(fechas) * 30))

    fig.update_layout(
        **base_layout(
            height,
            title=dict(text=f"<b>{titulo}</b>", x=0.01, xanchor="left"),
            margin=dict(l=125, r=30, t=58, b=75),
            xaxis=dict(
                title="Hora del turno",
                range=[0, 12],
                tickmode="array",
                tickvals=tickvals,
                ticktext=ticktext,
                gridcolor="#e5e7eb",
                zeroline=False,
                fixedrange=True,
            ),
            yaxis=dict(
                title="Fecha operativa · turno",
                tickmode="array",
                tickvals=y_tickvals,
                ticktext=y_ticktext,
                gridcolor="#eef2f7",
                zeroline=False,
                fixedrange=True,
                # Margen extra arriba para que la primera barra no quede recortada.
                range=[len(fechas) - 0.5, -0.85],
            ),
            legend=dict(orientation="h", y=-0.12, x=0),
            hovermode="closest",
        )
    )

    fig.add_vline(
        x=12,
        line_width=1.6,
        line_dash="dash",
        line_color="#b45309",
        annotation_text="Fin turno",
        annotation_position="top",
    )

    for x_ref in [2, 4, 6, 8, 10]:
        fig.add_vline(
            x=x_ref,
            line_width=0.9,
            line_dash="dot",
            line_color="#cbd5e1",
        )

    return fig





def preparar_primeros_inicios_distribucion(ciclos: pd.DataFrame) -> pd.DataFrame:
    """
    Conserva únicamente el primer inicio de cada Jumbo + fecha operativa + turno.
    X_Inicio está en horas relativas desde el inicio del turno (0..12), por lo
    que el turno noche se maneja correctamente aunque cruce medianoche.
    """
    if ciclos.empty:
        return pd.DataFrame()

    requeridas = {
        "Fecha_Operativa_DT",
        "Turno",
        "Jumbo",
        "X_Inicio",
    }
    if not requeridas.issubset(ciclos.columns):
        return pd.DataFrame()

    work = ciclos.copy()
    work["X_Inicio"] = pd.to_numeric(work["X_Inicio"], errors="coerce")
    work = work[
        work["Fecha_Operativa_DT"].notna()
        & work["Turno"].notna()
        & work["Jumbo"].notna()
        & work["X_Inicio"].notna()
        & work["X_Inicio"].between(0, 12, inclusive="both")
    ].copy()

    if work.empty:
        return pd.DataFrame()

    primeros = (
        work.sort_values(
            [
                "Fecha_Operativa_DT",
                "Turno",
                "Jumbo",
                "X_Inicio",
                "Ciclo",
            ]
        )
        .drop_duplicates(
            subset=[
                "Fecha_Operativa_DT",
                "Turno",
                "Jumbo",
            ],
            keep="first",
        )
        .copy()
    )

    return primeros.reset_index(drop=True)


def _densidad_gaussiana_horas(valores, grid):
    """
    KDE gaussiana simple sin scipy.
    El ancho de banda usa una adaptación de Silverman y se limita para evitar
    curvas demasiado irregulares cuando existen pocos ciclos.
    """
    vals = np.asarray(valores, dtype=float)
    vals = vals[np.isfinite(vals)]
    grid = np.asarray(grid, dtype=float)

    if vals.size == 0:
        return np.zeros_like(grid, dtype=float)

    if vals.size == 1:
        bandwidth = 0.40
    else:
        std = float(np.std(vals, ddof=1))
        q75, q25 = np.percentile(vals, [75, 25])
        iqr_sigma = float((q75 - q25) / 1.349) if q75 > q25 else std
        escala = min(std, iqr_sigma) if std > 0 and iqr_sigma > 0 else max(std, iqr_sigma)
        if not np.isfinite(escala) or escala <= 0:
            escala = 0.50
        bandwidth = 0.9 * escala * (vals.size ** (-1 / 5))
        bandwidth = float(np.clip(bandwidth, 0.25, 0.85))

    z = (grid[:, None] - vals[None, :]) / bandwidth
    densidad = np.exp(-0.5 * z * z).sum(axis=1)
    densidad /= vals.size * bandwidth * np.sqrt(2 * np.pi)
    return densidad


def grafico_distribucion_primeros_inicios(
    primeros: pd.DataFrame,
    turno: str,
):
    """
    Histograma de primeros inicios + curva de densidad para un turno.

    Histograma: frecuencia en intervalos de 30 min.
    Curva: KDE gaussiana en eje Y secundario.
    Incluye Promedio, Mediana y Pico aproximado de la densidad.
    """
    if primeros.empty:
        return None

    g = primeros[primeros["Turno"].eq(turno)].copy()
    if g.empty:
        return None

    valores = pd.to_numeric(g["X_Inicio"], errors="coerce").dropna().to_numpy(dtype=float)
    valores = valores[(valores >= 0) & (valores <= 12)]
    if len(valores) == 0:
        return None

    # Intervalos de 30 minutos.
    edges = np.arange(0, 12.0001 + 0.5, 0.5)
    counts, _ = np.histogram(valores, bins=edges)
    centers = (edges[:-1] + edges[1:]) / 2

    # Curva de densidad suave.
    grid = np.linspace(0, 12, 241)
    densidad = _densidad_gaussiana_horas(valores, grid)

    promedio = float(np.mean(valores))
    mediana = float(np.median(valores))
    pico = float(grid[int(np.argmax(densidad))]) if len(densidad) else mediana

    fig = make_subplots(specs=[[{"secondary_y": True}]])

    fig.add_trace(
        go.Bar(
            x=centers,
            y=counts,
            width=0.46,
            name="Frecuencia · 30 min",
            marker=dict(
                color="rgba(100,116,139,0.42)",
                line=dict(color="#64748b", width=0.8),
            ),
            customdata=[
                [
                    _hora_relativa_turno_a_texto(max(0, c - 0.25), turno),
                    _hora_relativa_turno_a_texto(min(12, c + 0.25), turno),
                    int(n),
                ]
                for c, n in zip(centers, counts)
            ],
            hovertemplate=(
                "<b>Primeros inicios</b>"
                "<br>Rango: %{customdata[0]}–%{customdata[1]}"
                "<br>Cantidad: %{customdata[2]}"
                "<extra></extra>"
            ),
        ),
        secondary_y=False,
    )

    fig.add_trace(
        go.Scatter(
            x=grid,
            y=densidad,
            mode="lines",
            name="Densidad suavizada",
            line=dict(color="#0f766e", width=3),
            hovertemplate=(
                "<b>Densidad</b>"
                "<br>Hora: %{customdata}"
                "<br>Densidad: %{y:.3f}"
                "<extra></extra>"
            ),
            customdata=[_hora_relativa_turno_a_texto(v, turno) for v in grid],
        ),
        secondary_y=True,
    )

    # Promedio y mediana.
    fig.add_vline(
        x=promedio,
        line_width=1.6,
        line_dash="dash",
        line_color="#111827",
        annotation_text=f"Promedio {_hora_relativa_turno_a_texto(promedio, turno)}",
        annotation_position="top right",
    )
    fig.add_vline(
        x=mediana,
        line_width=1.8,
        line_dash="dot",
        line_color="#2563eb",
        annotation_text=f"Mediana {_hora_relativa_turno_a_texto(mediana, turno)}",
        annotation_position="top left",
    )

    # Pico aproximado de la densidad.
    fig.add_vline(
        x=pico,
        line_width=1.1,
        line_dash="dot",
        line_color="#0f766e",
        opacity=0.65,
    )

    tickvals = list(range(0, 13))
    if turno == "Día":
        ticktext = [f"{(7 + h) % 24:02d}:00" for h in tickvals]
        titulo = "Turno Día · Distribución de primeros inicios"
    else:
        ticktext = [f"{(19 + h) % 24:02d}:00" for h in tickvals]
        titulo = "Turno Noche · Distribución de primeros inicios"

    equipos = int(g["Jumbo"].nunique())
    observaciones = int(len(g))

    fig.update_layout(
        **base_layout(
            480,
            title=dict(
                text=f"<b>{titulo}</b>",
                x=0.01,
                xanchor="left",
            ),
            margin=dict(l=70, r=70, t=80, b=70),
            xaxis=dict(
                title="Hora del primer inicio",
                range=[0, 12],
                tickmode="array",
                tickvals=tickvals,
                ticktext=ticktext,
                fixedrange=True,
                gridcolor="#e5e7eb",
                zeroline=False,
            ),
            legend=dict(
                orientation="h",
                y=-0.18,
                x=0,
            ),
            hovermode="closest",
        )
    )

    fig.update_yaxes(
        title_text="Cantidad de primeros inicios",
        rangemode="tozero",
        fixedrange=True,
        secondary_y=False,
    )
    fig.update_yaxes(
        title_text="Densidad",
        rangemode="tozero",
        fixedrange=True,
        showgrid=False,
        secondary_y=True,
    )

    # Nota compacta dentro del gráfico.
    fig.add_annotation(
        x=0.995,
        y=1.11,
        xref="paper",
        yref="paper",
        xanchor="right",
        yanchor="top",
        showarrow=False,
        align="right",
        text=(
            f"Pico aprox.: <b>{_hora_relativa_turno_a_texto(pico, turno)}</b> · "
            f"{observaciones} inicios · {equipos} equipo(s)"
        ),
        font=dict(size=11, color="#475569"),
    )

    return fig


def preparar_tendencia_inicio_diario(ciclos: pd.DataFrame) -> pd.DataFrame:
    """
    Construye un valor diario consolidado de primer inicio por turno.

    1) Para cada Fecha operativa + Turno + Jumbo conserva únicamente
       el primer inicio del equipo.
    2) Para cada Fecha operativa + Turno calcula:
       - Promedio de los primeros inicios de los equipos disponibles.
       - Mediana de los primeros inicios de los equipos disponibles.

    X_Inicio está expresado como horas transcurridas desde el inicio
    del turno, por lo que el cálculo del turno noche es continuo y no
    se distorsiona al cruzar medianoche.
    """
    if ciclos.empty:
        return pd.DataFrame()

    requeridas = {
        "Fecha_Operativa_DT",
        "Turno",
        "Jumbo",
        "X_Inicio",
    }
    if not requeridas.issubset(ciclos.columns):
        return pd.DataFrame()

    work = ciclos.copy()
    work["X_Inicio"] = pd.to_numeric(
        work["X_Inicio"],
        errors="coerce",
    )
    work = work[
        work["Fecha_Operativa_DT"].notna()
        & work["Turno"].notna()
        & work["Jumbo"].notna()
        & work["X_Inicio"].notna()
    ].copy()

    if work.empty:
        return pd.DataFrame()

    # Primer inicio real de cada equipo en cada fecha-turno.
    primeros = (
        work.sort_values(
            [
                "Fecha_Operativa_DT",
                "Turno",
                "Jumbo",
                "X_Inicio",
                "Ciclo",
            ]
        )
        .drop_duplicates(
            subset=[
                "Fecha_Operativa_DT",
                "Turno",
                "Jumbo",
            ],
            keep="first",
        )
        .copy()
    )

    resumen = (
        primeros.groupby(
            ["Fecha_Operativa_DT", "Turno"],
            as_index=False,
        )
        .agg(
            Promedio_h=("X_Inicio", "mean"),
            Mediana_h=("X_Inicio", "median"),
            Equipos=("Jumbo", "nunique"),
        )
        .sort_values(["Turno", "Fecha_Operativa_DT"])
        .reset_index(drop=True)
    )

    return resumen


def _hora_relativa_turno_a_texto(valor_h, turno: str) -> str:
    """Convierte hora relativa del turno a HH:MM."""
    if valor_h is None or pd.isna(valor_h):
        return "-"

    minutos = int(round(float(valor_h) * 60))
    inicio_h = 7 if turno == "Día" else 19
    total_min = (inicio_h * 60 + minutos) % (24 * 60)
    hh = total_min // 60
    mm = total_min % 60
    return f"{hh:02d}:{mm:02d}"


def grafico_tendencia_inicio_diario(
    resumen: pd.DataFrame,
    turno: str,
    mostrar_promedio: bool = True,
    mostrar_mediana: bool = True,
):
    """
    Curva diaria de la hora consolidada del primer inicio.

    Eje X: fecha operativa.
    Eje Y: hora dentro del turno (0 a 12 h desde el inicio).
    """
    if resumen.empty:
        return None

    g = resumen[
        resumen["Turno"].eq(turno)
    ].copy()

    if g.empty or (not mostrar_promedio and not mostrar_mediana):
        return None

    g = g.sort_values("Fecha_Operativa_DT")

    fig = go.Figure()

    fechas_txt = [
        pd.Timestamp(x).strftime("%d/%m/%Y")
        for x in g["Fecha_Operativa_DT"]
    ]

    if mostrar_promedio:
        horas_prom = [
            _hora_relativa_turno_a_texto(v, turno)
            for v in g["Promedio_h"]
        ]
        custom_prom = list(
            zip(
                fechas_txt,
                horas_prom,
                g["Equipos"].astype(int).tolist(),
            )
        )

        fig.add_trace(
            go.Scatter(
                x=g["Fecha_Operativa_DT"],
                y=g["Promedio_h"],
                mode="lines+markers",
                name="Promedio",
                line=dict(
                    color="#111827",
                    width=2.4,
                    dash="solid",
                ),
                marker=dict(
                    size=7,
                    symbol="circle",
                    color="#111827",
                ),
                customdata=custom_prom,
                hovertemplate=(
                    "<b>Promedio</b>"
                    "<br>Fecha: %{customdata[0]}"
                    "<br>Hora: %{customdata[1]}"
                    "<br>Equipos considerados: %{customdata[2]}"
                    "<extra></extra>"
                ),
            )
        )

    if mostrar_mediana:
        horas_med = [
            _hora_relativa_turno_a_texto(v, turno)
            for v in g["Mediana_h"]
        ]
        custom_med = list(
            zip(
                fechas_txt,
                horas_med,
                g["Equipos"].astype(int).tolist(),
            )
        )

        # La mediana se dibuja después del promedio y con línea punteada.
        # Con 2 equipos ambos valores coinciden; el patrón punteado permite
        # reconocer que ambas series están superpuestas.
        fig.add_trace(
            go.Scatter(
                x=g["Fecha_Operativa_DT"],
                y=g["Mediana_h"],
                mode="lines+markers",
                name="Mediana",
                line=dict(
                    color="#64748b",
                    width=2.4,
                    dash="dash",
                ),
                marker=dict(
                    size=8,
                    symbol="square-open",
                    color="#64748b",
                    line=dict(
                        color="#64748b",
                        width=1.5,
                    ),
                ),
                customdata=custom_med,
                hovertemplate=(
                    "<b>Mediana</b>"
                    "<br>Fecha: %{customdata[0]}"
                    "<br>Hora: %{customdata[1]}"
                    "<br>Equipos considerados: %{customdata[2]}"
                    "<extra></extra>"
                ),
            )
        )

    tickvals_y = list(range(0, 13))
    if turno == "Día":
        ticktext_y = [
            f"{(7 + h) % 24:02d}:00"
            for h in tickvals_y
        ]
        titulo = "Turno Día · Tendencia del primer inicio consolidado"
    else:
        ticktext_y = [
            f"{(19 + h) % 24:02d}:00"
            for h in tickvals_y
        ]
        titulo = "Turno Noche · Tendencia del primer inicio consolidado"

    fechas = g["Fecha_Operativa_DT"].tolist()
    tickvals_x = fechas[::2] if len(fechas) > 12 else fechas
    ticktext_x = [
        pd.Timestamp(x).strftime("%d/%m")
        for x in tickvals_x
    ]

    fig.update_layout(
        **base_layout(
            460,
            title=dict(
                text=f"<b>{titulo}</b>",
                x=0.01,
                xanchor="left",
            ),
            margin=dict(l=90, r=30, t=60, b=80),
            xaxis=dict(
                title="Fecha operativa",
                tickmode="array",
                tickvals=tickvals_x,
                ticktext=ticktext_x,
                tickangle=-35,
                gridcolor="#eef2f7",
                fixedrange=True,
            ),
            yaxis=dict(
                title="Hora de primer inicio",
                range=[0, 12],
                tickmode="array",
                tickvals=tickvals_y,
                ticktext=ticktext_y,
                gridcolor="#e5e7eb",
                zeroline=False,
                fixedrange=True,
            ),
            legend=dict(
                orientation="h",
                y=1.09,
                x=0.70,
                xanchor="left",
            ),
            hovermode="x unified",
        )
    )

    # Guías cada 2 horas para mantener consistencia con el timeline.
    for y_ref in [2, 4, 6, 8, 10]:
        fig.add_hline(
            y=y_ref,
            line_width=0.8,
            line_dash="dot",
            line_color="#cbd5e1",
            layer="below",
        )

    return fig


def preparar_tendencia_ultimo_fin_diario(ciclos: pd.DataFrame) -> pd.DataFrame:
    """
    Construye un valor diario consolidado de la hora de término del último
    ciclo/round por turno.

    1) Para cada Fecha operativa + Turno + Jumbo conserva el ciclo cuyo
       X_Fin sea más tardío dentro de ese turno. Si un equipo realizó dos
       o más rounds, se toma el Fin del último round ejecutado.
    2) Para cada Fecha operativa + Turno calcula:
       - Promedio de las últimas horas de término de los equipos disponibles.
       - Mediana de las últimas horas de término de los equipos disponibles.

    X_Fin está expresado como horas transcurridas desde el inicio del turno,
    por lo que el cálculo del turno noche permanece continuo al cruzar
    medianoche.
    """
    if ciclos.empty:
        return pd.DataFrame()

    requeridas = {
        "Fecha_Operativa_DT",
        "Turno",
        "Jumbo",
        "X_Fin",
    }
    if not requeridas.issubset(ciclos.columns):
        return pd.DataFrame()

    work = ciclos.copy()
    work["X_Fin"] = pd.to_numeric(
        work["X_Fin"],
        errors="coerce",
    )
    work = work[
        work["Fecha_Operativa_DT"].notna()
        & work["Turno"].notna()
        & work["Jumbo"].notna()
        & work["X_Fin"].notna()
    ].copy()

    if work.empty:
        return pd.DataFrame()

    # Último término real de cada equipo en cada fecha-turno.
    ultimos = (
        work.sort_values(
            [
                "Fecha_Operativa_DT",
                "Turno",
                "Jumbo",
                "X_Fin",
                "Ciclo",
            ],
            ascending=[True, True, True, False, False],
        )
        .drop_duplicates(
            subset=[
                "Fecha_Operativa_DT",
                "Turno",
                "Jumbo",
            ],
            keep="first",
        )
        .copy()
    )

    resumen = (
        ultimos.groupby(
            ["Fecha_Operativa_DT", "Turno"],
            as_index=False,
        )
        .agg(
            Promedio_h=("X_Fin", "mean"),
            Mediana_h=("X_Fin", "median"),
            Equipos=("Jumbo", "nunique"),
        )
        .sort_values(["Turno", "Fecha_Operativa_DT"])
        .reset_index(drop=True)
    )

    return resumen


def grafico_tendencia_ultimo_fin_diario(
    resumen: pd.DataFrame,
    turno: str,
    mostrar_promedio: bool = True,
    mostrar_mediana: bool = True,
):
    """
    Curva diaria de la hora consolidada de término del último ciclo.

    Eje X: fecha operativa.
    Eje Y: hora dentro del turno medida desde el inicio del turno.
    """
    if resumen.empty:
        return None

    g = resumen[
        resumen["Turno"].eq(turno)
    ].copy()

    if g.empty or (not mostrar_promedio and not mostrar_mediana):
        return None

    g = g.sort_values("Fecha_Operativa_DT")

    fig = go.Figure()

    fechas_txt = [
        pd.Timestamp(x).strftime("%d/%m/%Y")
        for x in g["Fecha_Operativa_DT"]
    ]

    if mostrar_promedio:
        horas_prom = [
            _hora_relativa_turno_a_texto(v, turno)
            for v in g["Promedio_h"]
        ]
        custom_prom = list(
            zip(
                fechas_txt,
                horas_prom,
                g["Equipos"].astype(int).tolist(),
            )
        )

        fig.add_trace(
            go.Scatter(
                x=g["Fecha_Operativa_DT"],
                y=g["Promedio_h"],
                mode="lines+markers",
                name="Promedio",
                line=dict(
                    color="#111827",
                    width=2.4,
                    dash="solid",
                ),
                marker=dict(
                    size=7,
                    symbol="circle",
                    color="#111827",
                ),
                customdata=custom_prom,
                hovertemplate=(
                    "<b>Promedio</b>"
                    "<br>Fecha: %{customdata[0]}"
                    "<br>Último término: %{customdata[1]}"
                    "<br>Equipos considerados: %{customdata[2]}"
                    "<extra></extra>"
                ),
            )
        )

    if mostrar_mediana:
        horas_med = [
            _hora_relativa_turno_a_texto(v, turno)
            for v in g["Mediana_h"]
        ]
        custom_med = list(
            zip(
                fechas_txt,
                horas_med,
                g["Equipos"].astype(int).tolist(),
            )
        )

        fig.add_trace(
            go.Scatter(
                x=g["Fecha_Operativa_DT"],
                y=g["Mediana_h"],
                mode="lines+markers",
                name="Mediana",
                line=dict(
                    color="#64748b",
                    width=2.4,
                    dash="dash",
                ),
                marker=dict(
                    size=8,
                    symbol="square-open",
                    color="#64748b",
                    line=dict(
                        color="#64748b",
                        width=1.5,
                    ),
                ),
                customdata=custom_med,
                hovertemplate=(
                    "<b>Mediana</b>"
                    "<br>Fecha: %{customdata[0]}"
                    "<br>Último término: %{customdata[1]}"
                    "<br>Equipos considerados: %{customdata[2]}"
                    "<extra></extra>"
                ),
            )
        )

    # Mantener referencia del turno completo, pero permitir visualizar
    # ciclos que terminen ligeramente después de las 12 horas.
    max_val = pd.to_numeric(
        g[["Promedio_h", "Mediana_h"]].stack(),
        errors="coerce",
    ).max()
    y_max = max(12.0, float(max_val) if pd.notna(max_val) else 12.0)
    y_max = min(max(12.0, y_max + 0.25), 16.0)

    tickvals_y = list(range(0, int(y_max) + 1))
    if turno == "Día":
        ticktext_y = [
            f"{(7 + h) % 24:02d}:00"
            for h in tickvals_y
        ]
        titulo = "Turno Día · Tendencia del término del último ciclo"
    else:
        ticktext_y = [
            f"{(19 + h) % 24:02d}:00"
            for h in tickvals_y
        ]
        titulo = "Turno Noche · Tendencia del término del último ciclo"

    fechas = g["Fecha_Operativa_DT"].tolist()
    tickvals_x = fechas[::2] if len(fechas) > 12 else fechas
    ticktext_x = [
        pd.Timestamp(x).strftime("%d/%m")
        for x in tickvals_x
    ]

    fig.update_layout(
        **base_layout(
            460,
            title=dict(
                text=f"<b>{titulo}</b>",
                x=0.01,
                xanchor="left",
            ),
            margin=dict(l=90, r=30, t=60, b=80),
            xaxis=dict(
                title="Fecha operativa",
                tickmode="array",
                tickvals=tickvals_x,
                ticktext=ticktext_x,
                tickangle=-35,
                gridcolor="#eef2f7",
                fixedrange=True,
            ),
            yaxis=dict(
                title="Hora de término del último ciclo",
                range=[0, y_max],
                tickmode="array",
                tickvals=tickvals_y,
                ticktext=ticktext_y,
                gridcolor="#e5e7eb",
                zeroline=False,
                fixedrange=True,
            ),
            legend=dict(
                orientation="h",
                y=1.09,
                x=0.70,
                xanchor="left",
            ),
            hovermode="x unified",
        )
    )

    # Línea de referencia del fin nominal del turno.
    fig.add_hline(
        y=12,
        line_width=1.4,
        line_dash="dash",
        line_color="#b45309",
        annotation_text="Fin turno",
        annotation_position="top right",
    )

    for y_ref in [2, 4, 6, 8, 10]:
        fig.add_hline(
            y=y_ref,
            line_width=0.8,
            line_dash="dot",
            line_color="#cbd5e1",
            layer="below",
        )

    return fig


# ==========================================================
# PROCESAMIENTO MASIVO / CACHE
# ==========================================================


def guardar_resultado_desde_disco(item):
    """Procesa un archivo ya liberado del uploader y conserva solo resultados."""
    clave = item["clave"]
    path = Path(item["path"])
    try:
        resultado = procesar_archivo(
            path,
            nombre_archivo=item["nombre"],
            generar_visuales=False,
        )

        # Visuales diferidos: nunca se crean durante el lote masivo.
        fig = resultado.pop("fig", None)
        if fig is not None:
            plt.close(fig)
            del fig
        resultado.pop("plano_nav_png", None)
        resultado.pop("png_bytes", None)

        resultado["nombre_archivo"] = item["nombre"]
        resultado["error"] = None
        resultado["_cache_key"] = clave
        resultado["_source_path"] = str(path)

        st.session_state.procesados[clave] = resultado
    except Exception as exc:
        st.session_state.procesados[clave] = {
            "nombre_archivo": item["nombre"],
            "_cache_key": clave,
            "_source_path": str(path),
            "error": str(exc),
        }


def _visual_paths(cache_key):
    visual_dir = _session_work_dir() / "visuals"
    return (
        visual_dir / f"{cache_key}_boxplot.png",
        visual_dir / f"{cache_key}_plano.png",
    )


def asegurar_visuales_resultado(r):
    """Genera boxplot/plano solo cuando el usuario solicita ese ciclo."""
    cache_key = r.get("_cache_key") or hashlib.sha1(
        str(r.get("nombre_archivo", "")).encode("utf-8")
    ).hexdigest()
    box_path, nav_path = _visual_paths(cache_key)

    detalle = r.get("detalle")
    rep = r.get("resumen_reporte") or {}
    metadata = r.get("metadata") or rep
    fuente = str(rep.get("Fuente") or r.get("fuente") or "PDF").upper()
    source_path = Path(r.get("_source_path") or "")

    if not box_path.exists() and isinstance(detalle, pd.DataFrame) and not detalle.empty:
        fig = generar_grafico(detalle, metadata)
        fig.savefig(box_path, format="png", dpi=170, bbox_inches="tight", facecolor="white")
        plt.close(fig)
        del fig
        gc.collect()

    if not nav_path.exists():
        nav_bytes = None
        if fuente == "ZDA" and isinstance(detalle, pd.DataFrame) and not detalle.empty:
            nav_bytes = generar_plano_zda_png(detalle, metadata, resolution=150)
        elif fuente == "PDF" and source_path.exists():
            nav_bytes = extraer_plano_navegacion_png(source_path, resolution=150)

        if nav_bytes:
            nav_path.write_bytes(nav_bytes)
            del nav_bytes
            gc.collect()

    return box_path if box_path.exists() else None, nav_path if nav_path.exists() else None


# ==========================================================
# CARGA MASIVA EN DOS FASES
# ==========================================================


def _streamlit_version_tuple():
    """Versión numérica simple para habilitar carga de carpetas."""
    nums = re.findall(r"\d+", str(getattr(st, "__version__", "0.0.0")))
    vals = [int(x) for x in nums[:3]]
    while len(vals) < 3:
        vals.append(0)
    return tuple(vals)


with st.container(border=True):
    st.subheader("Cargar ciclos (.ZDA / PDF)")
    st.caption(
        "Puedes agregar archivos individuales o seleccionar una carpeta completa. "
        "Los ciclos ya procesados permanecen cargados y puedes seguir incorporando "
        "nuevos archivos a medida que avanza el mes."
    )

    col_files, col_folder, col_clear = st.columns([1.15, 1.15, 1.45])

    with col_files:
        with st.popover(
            "📄 Elegir archivos",
            use_container_width=True,
            help="Selecciona uno o varios archivos PDF/ZDA.",
        ):
            archivos_individuales = st.file_uploader(
                "Archivos PDF / ZDA",
                type=["pdf", "zda"],
                accept_multiple_files=True,
                key=f"uploader_archivos_{st.session_state.uploader_version}",
                label_visibility="collapsed",
                help=(
                    "Puedes seleccionar uno o varios archivos. "
                    "La carga se procesa automáticamente."
                ),
            )

    with col_folder:
        if _streamlit_version_tuple() >= (1, 49, 0):
            with st.popover(
                "📁 Elegir carpeta",
                use_container_width=True,
                help="Carga todos los PDF/ZDA contenidos en una carpeta y sus subcarpetas.",
            ):
                archivos_carpeta = st.file_uploader(
                    "Carpeta con archivos PDF / ZDA",
                    type=["pdf", "zda"],
                    accept_multiple_files="directory",
                    key=f"uploader_carpeta_{st.session_state.uploader_version}",
                    label_visibility="collapsed",
                    help=(
                        "Selecciona una carpeta. Solo se cargarán archivos PDF/ZDA; "
                        "también se consideran sus subcarpetas."
                    ),
                )
        else:
            archivos_carpeta = []
            st.button(
                "📁 Elegir carpeta",
                width="stretch",
                disabled=True,
                help="Requiere Streamlit 1.49 o superior.",
                key="folder_upload_disabled",
            )

    with col_clear:
        st.button(
            "🗑️ Borrar datos cargados",
            width="stretch",
            on_click=limpiar_analisis,
            help=(
                "Elimina de la sesión todos los ciclos procesados, "
                "archivos temporales y filtros asociados."
            ),
        )

    archivos_individuales = archivos_individuales or []
    archivos_carpeta = archivos_carpeta or []
    seleccion = list(archivos_individuales) + list(archivos_carpeta)

    total_upload_mb = sum(
        int(getattr(a, "size", 0) or 0)
        for a in seleccion
    ) / (1024 * 1024)

    procesados_ok_actual = sum(
        1
        for r in st.session_state.procesados.values()
        if not r.get("error")
    )
    procesados_error_actual = sum(
        1
        for r in st.session_state.procesados.values()
        if r.get("error")
    )

    estado_carga = (
        f"{procesados_ok_actual} ciclo(s) cargado(s) correctamente."
    )
    if procesados_error_actual:
        estado_carga += f" · {procesados_error_actual} archivo(s) con error."
    if seleccion:
        estado_carga += (
            f" · Nuevos seleccionados: {len(seleccion)} "
            f"({total_upload_mb:,.1f} MB)."
        )

    st.markdown(
        f"<div style='font-size:0.98rem; color:#475467; padding-top:0.25rem;'>"
        f"{estado_carga}</div>",
        unsafe_allow_html=True,
    )

if seleccion and total_upload_mb >= 700:
    st.warning(
        "El lote seleccionado supera aproximadamente 700 MB. El modo masivo "
        "libera los archivos antes del parsing, pero el navegador debe transferir "
        "primero todo el lote. Si la carga supera la memoria disponible, conviene "
        "dividir únicamente la selección en dos carpetas o lotes."
    )

# Primera fase automática: al seleccionar archivos o una carpeta, se copian
# a disco temporal y se libera inmediatamente el uploader antes del parsing.
if seleccion:
    estado_stage = st.empty()
    estado_stage.write(
        f"Preparando {len(seleccion)} archivo(s) en disco temporal..."
    )
    nuevos = preparar_archivos_en_disco(seleccion)
    estado_stage.empty()

    # Cambiar la versión de ambos uploaders hace que Streamlit libere sus bytes
    # antes de comenzar el procesamiento intensivo del lote.
    st.session_state.uploader_version += 1
    st.session_state.auto_process_staged = bool(nuevos)
    st.rerun()

# Segunda fase: ya sin UploadedFile en RAM.
if st.session_state.auto_process_staged and st.session_state.staged_queue:
    cola = list(st.session_state.staged_queue)
    barra = st.progress(0)
    estado = st.empty()

    for i, item in enumerate(cola, start=1):
        estado.write(f"Procesando {i}/{len(cola)}: {item['nombre']}")
        guardar_resultado_desde_disco(item)
        st.session_state.staged_queue = [
            x for x in st.session_state.staged_queue
            if x.get("clave") != item.get("clave")
        ]
        gc.collect()
        barra.progress(i / len(cola))

    barra.empty()
    estado.empty()
    st.session_state.auto_process_staged = False
    st.rerun()

resultados_validos = [
    r
    for r in st.session_state.procesados.values()
    if not r.get("error")
]
errores = [
    r
    for r in st.session_state.procesados.values()
    if r.get("error")
]

if not resultados_validos:
    if errores:
        st.error(
            "No hay archivos procesados correctamente todavía."
        )
        st.dataframe(
            pd.DataFrame([
                {
                    "Archivo": r.get("nombre_archivo"),
                    "Error": r.get("error"),
                }
                for r in errores
            ]),
            width="stretch",
            hide_index=True,
        )
    else:
        st.write(
            "Selecciona uno o varios archivos o una carpeta completa para "
            "comenzar el análisis."
        )
    st.stop()


# ==========================================================
# CONSOLIDACIÓN
# ==========================================================

report_rows = [dict(r["resumen_reporte"]) for r in resultados_validos]
df_reportes = pd.DataFrame(report_rows)

# Fuerza la clasificación V33 para PDF y ZDA con el mismo criterio.
df_reportes["Tipo_Disparo"] = df_reportes["Barrenos_Realizados"].apply(clasificar_tipo_disparo_v33)
df_reportes["Tipo_Roca"] = df_reportes["Plan_Perforacion"].apply(tipo_roca_desde_plan_texto)

def _operador_filtro_row(r):
    for campo in ("Operador_ZDA", "Operador", "Operario"):
        valor = r.get(campo)
        if valor is not None and not pd.isna(valor):
            texto = str(valor).strip()
            if texto:
                return texto
    return "SIN DATO"

df_reportes["Operador_Filtro"] = df_reportes.apply(_operador_filtro_row, axis=1)
df_reportes["Considerado_KPI_Automatizacion"] = df_reportes["Tipo_Disparo"].eq("FRENTE")
for r in resultados_validos:
    rr = r["resumen_reporte"]
    rr["Tipo_Disparo"] = clasificar_tipo_disparo_v33(rr.get("Barrenos_Realizados"))
    rr["Tipo_Roca"] = tipo_roca_desde_plan_texto(rr.get("Plan_Perforacion"))
    rr["Operador_Filtro"] = (
        rr.get("Operador_ZDA")
        or rr.get("Operador")
        or rr.get("Operario")
        or "SIN DATO"
    )
    rr["Considerado_KPI_Automatizacion"] = rr["Tipo_Disparo"] == "FRENTE"

# HTML V33 solo agrega Resumen_Ciclos de reportes cuyo conteo está OK.
df_resumen = concatenar_dataframes(resultados_validos, "resumen_ciclo", solo_ok=True)
df_detalle = concatenar_dataframes(resultados_validos, "detalle")
df_atipicos = concatenar_dataframes(resultados_validos, "atipicos")
df_automatico = df_reportes.copy()

df_zda = df_reportes[df_reportes["Fuente"].eq("ZDA")].copy() if "Fuente" in df_reportes.columns else pd.DataFrame()

# ==========================================================
# FILTRO GLOBAL DE FECHAS EN SIDEBAR
# ==========================================================
# El contenedor fue creado al inicio dentro de st.sidebar para conservar
# la ubicación del filtro junto con Jumbos / Tipo / Roca / Operadores.
# Se llena aquí, cuando df_zda ya está consolidado, independientemente
# de qué sección del dashboard esté seleccionada.

global_fecha_inicio_zda = None
global_fecha_fin_zda = None

if sidebar_fecha_container is not None:
    with sidebar_fecha_container:
        st.markdown("##### Rango de fechas")
        st.caption("Aplica a Uso Automático y Tiempos de Ciclo.")

        if (
            not df_zda.empty
            and "Inicio_Perforacion_TS" in df_zda.columns
            and df_zda["Inicio_Perforacion_TS"].notna().any()
        ):
            fechas_sidebar = (
                df_zda.loc[
                    df_zda["Inicio_Perforacion_TS"].notna(),
                    "Inicio_Perforacion_TS",
                ]
                .apply(zda_operational_date)
            )

            fechas_sidebar = pd.to_datetime(
                fechas_sidebar,
                errors="coerce",
                utc=True,
            ).dropna()

            if not fechas_sidebar.empty:
                fecha_min_sidebar = fechas_sidebar.min().date()
                fecha_max_sidebar = fechas_sidebar.max().date()

                # Recuperar valores previos y ajustarlos al rango disponible.
                fecha_inicio_previa = st.session_state.get(
                    "fecha_inicio_zda_global",
                    fecha_min_sidebar,
                )
                fecha_fin_previa = st.session_state.get(
                    "fecha_fin_zda_global",
                    fecha_max_sidebar,
                )

                try:
                    fecha_inicio_previa = max(
                        fecha_min_sidebar,
                        min(fecha_inicio_previa, fecha_max_sidebar),
                    )
                except Exception:
                    fecha_inicio_previa = fecha_min_sidebar

                try:
                    fecha_fin_previa = max(
                        fecha_min_sidebar,
                        min(fecha_fin_previa, fecha_max_sidebar),
                    )
                except Exception:
                    fecha_fin_previa = fecha_max_sidebar

                # Si el rango previo queda invertido, volver al rango completo.
                if fecha_inicio_previa > fecha_fin_previa:
                    fecha_inicio_previa = fecha_min_sidebar
                    fecha_fin_previa = fecha_max_sidebar

                global_fecha_inicio_zda = st.date_input(
                    "Fecha inicio",
                    value=fecha_inicio_previa,
                    min_value=fecha_min_sidebar,
                    max_value=fecha_max_sidebar,
                    key="fecha_inicio_zda_global",
                    format="DD/MM/YYYY",
                )

                global_fecha_fin_zda = st.date_input(
                    "Fecha fin",
                    value=fecha_fin_previa,
                    min_value=fecha_min_sidebar,
                    max_value=fecha_max_sidebar,
                    key="fecha_fin_zda_global",
                    format="DD/MM/YYYY",
                )

                st.caption(
                    f"Rango mostrado: "
                    f"{global_fecha_inicio_zda.strftime('%d/%m/%Y')} "
                    f"→ {global_fecha_fin_zda.strftime('%d/%m/%Y')}"
                )
            else:
                st.caption("Sin fechas ZDA válidas.")
        else:
            st.caption("Sin fechas ZDA disponibles.")


# ==========================================================
# EXPORTACIÓN EXCEL - BD-PERFO + Resumen_Reportes + Resumen_Ciclos
# ==========================================================

df_bd_perfo = construir_bd_perfo(df_reportes, df_detalle)
excel_bytes = crear_excel_publicacion(df_bd_perfo, df_reportes, df_resumen)
nombre_excel = f"EBR Drill Analytics {datetime.now().strftime('%d-%m-%Y')}.xlsx"

st.download_button(
    "Descargar Excel consolidado",
    data=excel_bytes,
    file_name=nombre_excel,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
)



def preparar_barrenos_por_brazo(
    df_detalle: pd.DataFrame,
    df_reportes: pd.DataFrame,
) -> pd.DataFrame:
    """
    Cuenta los barrenos realizados por Brazo 1 y Brazo 2 en cada ciclo.

    Si para un mismo Jumbo+Ciclo existen registros PDF y ZDA, se prioriza
    ZDA para evitar duplicar el mismo round.
    El conteo incluye todos los tipos de barreno disponibles en el detalle.
    """
    if df_detalle.empty:
        return pd.DataFrame()

    requeridas = {"Jumbo", "Ciclo", "Boom"}
    if not requeridas.issubset(df_detalle.columns):
        return pd.DataFrame()

    det = df_detalle.copy()
    det["Boom"] = pd.to_numeric(det["Boom"], errors="coerce")
    det = det[det["Boom"].isin([1, 2])].copy()
    if det.empty:
        return pd.DataFrame()

    # Clave de ciclo. Fecha ayuda a evitar colisiones si un número de ciclo
    # se reutilizara en fechas diferentes.
    keys = ["Jumbo", "Ciclo"]
    if "Fecha_Inicio" in det.columns:
        keys.append("Fecha_Inicio")

    # Priorizar ZDA sobre PDF si el mismo ciclo está presente en ambos.
    if "Fuente" in det.columns:
        det["_fuente_norm"] = det["Fuente"].fillna("").astype(str).str.upper()
        zda_keys = set(
            map(
                tuple,
                det.loc[det["_fuente_norm"].eq("ZDA"), keys]
                .astype(str)
                .to_numpy(),
            )
        )

        if zda_keys:
            tuples = list(map(tuple, det[keys].astype(str).to_numpy()))
            keep = [
                (t not in zda_keys) or (fuente == "ZDA")
                for t, fuente in zip(tuples, det["_fuente_norm"])
            ]
            det = det.loc[keep].copy()

    counts = (
        det.groupby(keys + ["Boom"], dropna=False)
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )

    if 1 not in counts.columns:
        counts[1] = 0
    if 2 not in counts.columns:
        counts[2] = 0

    counts = counts.rename(
        columns={
            1: "Barrenos_B1",
            2: "Barrenos_B2",
        }
    )
    counts["Barrenos_B1"] = pd.to_numeric(
        counts["Barrenos_B1"], errors="coerce"
    ).fillna(0).astype(int)
    counts["Barrenos_B2"] = pd.to_numeric(
        counts["Barrenos_B2"], errors="coerce"
    ).fillna(0).astype(int)
    counts["Total_Barrenos_Brazos"] = (
        counts["Barrenos_B1"] + counts["Barrenos_B2"]
    )

    # Participación porcentual por brazo dentro de cada ciclo.
    total_seguro = counts["Total_Barrenos_Brazos"].replace(0, pd.NA)
    counts["Pct_B1"] = (
        counts["Barrenos_B1"] / total_seguro * 100
    ).fillna(0.0)
    counts["Pct_B2"] = (
        counts["Barrenos_B2"] / total_seguro * 100
    ).fillna(0.0)

    # Añadir metadatos del ciclo para filtros y tooltip.
    if not df_reportes.empty:
        meta_cols = [
            c
            for c in [
                "Jumbo",
                "Ciclo",
                "Fecha_Inicio",
                "Hora_Inicio",
                "Tipo_Disparo",
                "Tipo_Roca",
                "Fuente",
            ]
            if c in df_reportes.columns
        ]
        meta = df_reportes[meta_cols].copy()

        # Prioridad ZDA en el metadata si hay PDF y ZDA del mismo ciclo.
        if "Fuente" in meta.columns:
            meta["_prioridad"] = (
                meta["Fuente"]
                .fillna("")
                .astype(str)
                .str.upper()
                .eq("ZDA")
                .astype(int)
            )
            meta = meta.sort_values("_prioridad", ascending=False)

        merge_keys = [
            c
            for c in ["Jumbo", "Ciclo", "Fecha_Inicio"]
            if c in counts.columns and c in meta.columns
        ]
        if not merge_keys:
            merge_keys = [
                c
                for c in ["Jumbo", "Ciclo"]
                if c in counts.columns and c in meta.columns
            ]

        meta = meta.drop_duplicates(subset=merge_keys, keep="first")
        meta = meta.drop(
            columns=["_prioridad"],
            errors="ignore",
        )

        # Evita duplicar columnas que ya vienen desde el detalle.
        add_cols = merge_keys + [
            c
            for c in [
                "Hora_Inicio",
                "Tipo_Disparo",
                "Tipo_Roca",
                "Fuente",
            ]
            if c in meta.columns and c not in merge_keys
        ]
        counts = counts.merge(
            meta[add_cols],
            on=merge_keys,
            how="left",
        )

    if "Tipo_Disparo" not in counts.columns:
        counts["Tipo_Disparo"] = "SIN CLASIFICAR"
    if "Tipo_Roca" not in counts.columns:
        counts["Tipo_Roca"] = "SIN DATO"

    return counts


def grafico_b1_participacion_unico(
    df_brazos: pd.DataFrame,
):
    """
    Gráfico único por ciclo con la participación del Brazo 1.

    - Eje Y: % de participación de B1 respecto al total B1+B2.
    - Si el valor es > 50%, B1 realizó más barrenos que B2.
    - Si el valor es < 50%, B2 realizó más barrenos que B1.
    - Una línea horizontal en 50% sirve como referencia visual.
    """
    if df_brazos.empty:
        return None

    g = df_brazos.copy()
    g = asegurar_fechahora(g)
    if g.empty:
        return None

    g["Ciclo_Label"] = g["Ciclo"].apply(
        lambda v: (
            f"C{int(v)}"
            if pd.notna(v)
            and str(v).replace(".", "", 1).isdigit()
            else f"C{v}"
        )
    )

    # Texto corto del estado de balance.
    def _dominancia(row):
        if float(row.get("Pct_B1", 0)) > 50:
            return "B1 > B2"
        if float(row.get("Pct_B1", 0)) < 50:
            return "B2 > B1"
        return "Balanceado"

    g["Dominancia_B1"] = g.apply(_dominancia, axis=1)

    # Orden cronológico y etiqueta compacta para un gráfico único.
    g = g.reset_index(drop=True)
    g["Orden_Ciclo"] = range(1, len(g) + 1)
    g["X_Label"] = g.apply(
        lambda r: f"{str(r['Jumbo'])[-3:]}·{r['Ciclo_Label']}",
        axis=1,
    )

    fig = go.Figure()
    colores = {
        "JUMB001": "#4f6df5",
        "JUMB002": "#f05a3b",
    }

    for jumbo in sorted(g["Jumbo"].dropna().astype(str).unique()):
        gj = g[g["Jumbo"].astype(str).eq(str(jumbo))].copy()
        if gj.empty:
            continue

        custom = gj[
            [
                "Jumbo",
                "Ciclo",
                "Barrenos_B1",
                "Barrenos_B2",
                "Total_Barrenos_Brazos",
                "Pct_B1",
                "Pct_B2",
                "Tipo_Roca",
                "Tipo_Disparo",
                "Fecha_Inicio",
                "Dominancia_B1",
            ]
        ].astype(object).to_numpy()

        fig.add_trace(
            go.Scatter(
                name=jumbo,
                x=gj["X_Label"],
                y=gj["Pct_B1"],
                mode="lines+markers",
                customdata=custom,
                line=dict(
                    width=2.5,
                    shape="spline",
                    smoothing=0.80,
                    color=colores.get(jumbo),
                ),
                marker=dict(size=8, color=colores.get(jumbo)),
                hovertemplate=(
                    "<b>%{customdata[0]} · Ciclo %{customdata[1]}</b><br>"
                    "Participación B1: %{customdata[5]:.1f}%<br>"
                    "Participación B2: %{customdata[6]:.1f}%<br>"
                    "Lectura visual: %{customdata[10]}<br>"
                    "Brazo 1: %{customdata[2]} barrenos<br>"
                    "Brazo 2: %{customdata[3]} barrenos<br>"
                    "Total: %{customdata[4]} barrenos<br>"
                    "Tipo de roca: %{customdata[7]}<br>"
                    "Tipo de disparo: %{customdata[8]}<br>"
                    "Fecha: %{customdata[9]}"
                    "<extra></extra>"
                ),
            )
        )

    fig.update_layout(
        **base_layout(
            480,
            margin=dict(l=72, r=30, t=70, b=95),
            xaxis=dict(
                title="Ciclo / round",
                type="category",
                categoryorder="array",
                categoryarray=g["X_Label"].tolist(),
                tickangle=-45,
                gridcolor="#eef2f7",
            ),
            yaxis=dict(
                title="Participación del Brazo 1 (%)",
                range=[0, 100],
                tickmode="array",
                tickvals=[0, 20, 40, 50, 60, 80, 100],
                ticktext=["0%", "20%", "40%", "50%", "60%", "80%", "100%"],
                gridcolor="#eef2f7",
            ),
            shapes=[
                dict(
                    type="line",
                    xref="paper",
                    x0=0,
                    x1=1,
                    yref="y",
                    y0=50,
                    y1=50,
                    line=dict(
                        color="#475569",
                        width=2.0,
                        dash="dot",
                    ),
                    layer="above",
                ),
            ],
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.02,
                xanchor="left",
                x=0,
            ),
        )
    )

    return fig






def aplicar_filtro_fechas_global(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aplica Fecha inicio / Fecha fin del sidebar a un DataFrame
    usando la columna Fecha_Inicio (dd/mm/YYYY).

    Si no hay rango seleccionado o no existe Fecha_Inicio,
    devuelve el DataFrame sin cambios.
    """
    if df is None or df.empty or "Fecha_Inicio" not in df.columns:
        return df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()

    fecha_inicio = st.session_state.get("fecha_inicio_zda_global")
    fecha_fin = st.session_state.get("fecha_fin_zda_global")

    if fecha_inicio is None or fecha_fin is None:
        return df.copy()

    fechas = pd.to_datetime(
        df["Fecha_Inicio"],
        format="%d/%m/%Y",
        errors="coerce",
    )

    mask = (
        fechas.notna()
        & (fechas.dt.date >= fecha_inicio)
        & (fechas.dt.date <= fecha_fin)
    )

    return df.loc[mask].copy()


def _nombre_jumbo_resumen(valor):
    texto = str(valor or "").strip()
    m = re.fullmatch(r"JUMB0*(\d+)", texto, re.IGNORECASE)
    if m:
        return f"Jumbo {int(m.group(1))}"
    return texto or "-"


def _fecha_corta_es(valor):
    if valor is None or pd.isna(valor):
        return "-"
    ts = pd.Timestamp(valor)
    meses = [
        "ene.", "feb.", "mar.", "abr.", "may.", "jun.",
        "jul.", "ago.", "sep.", "oct.", "nov.", "dic.",
    ]
    return f"{ts.day:02d}-{meses[ts.month - 1]}"


def render_kpis_uso_automatico(df_auto: pd.DataFrame):
    """
    Resumen superior de la sección Uso Automático.

    Los indicadores se recalculan con el rango Fecha inicio / Fecha fin
    seleccionado en el panel lateral.
    """
    if df_auto is None or df_auto.empty:
        return

    base = aplicar_filtro_fechas_global(df_auto)
    total_ciclos = len(base)

    # ------------------------------------------------------
    # Ciclos cargados + desglose por jumbo
    # ------------------------------------------------------
    if "Jumbo" in base.columns:
        vc = (
            base["Jumbo"]
            .fillna("SIN DATO")
            .astype(str)
            .value_counts()
        )

        def _sort_jumbo(item):
            nombre, _ = item
            m = re.fullmatch(r"JUMB0*(\d+)", str(nombre), re.IGNORECASE)
            if m:
                return (0, int(m.group(1)))
            return (1, str(nombre))

        jumbo_detalle = " · ".join(
            f"{_nombre_jumbo_resumen(jumbo)}: {int(n)}"
            for jumbo, n in sorted(vc.items(), key=_sort_jumbo)
        )
    else:
        jumbo_detalle = f"{total_ciclos} ciclos"

    # ------------------------------------------------------
    # Rango de fechas
    # ------------------------------------------------------
    fechas = pd.Series(dtype="datetime64[ns]")
    if "Fecha_Inicio" in base.columns:
        fechas = pd.to_datetime(
            base["Fecha_Inicio"],
            format="%d/%m/%Y",
            errors="coerce",
        ).dropna()

    fecha_inicio_sel = st.session_state.get("fecha_inicio_zda_global")
    fecha_fin_sel = st.session_state.get("fecha_fin_zda_global")

    if fecha_inicio_sel is not None and fecha_fin_sel is not None:
        rango_fecha = (
            f"{_fecha_corta_es(pd.Timestamp(fecha_inicio_sel))} – "
            f"{_fecha_corta_es(pd.Timestamp(fecha_fin_sel))}"
        )
        fechas_sub = f"{len(fechas)} ciclos en el rango"
    elif not fechas.empty:
        fecha_min = fechas.min()
        fecha_max = fechas.max()
        rango_fecha = f"{_fecha_corta_es(fecha_min)} – {_fecha_corta_es(fecha_max)}"
        fechas_sub = f"{len(fechas)} ciclos con fecha"
    else:
        rango_fecha = "-"
        fechas_sub = "Sin fechas válidas"

    # ------------------------------------------------------
    # Horas automáticas totales
    # ------------------------------------------------------
    auto_min = (
        pd.to_numeric(
            base.get(
                "Auto_Total_Brazos_min",
                pd.Series(index=base.index, dtype=float),
            ),
            errors="coerce",
        )
    )
    manual_min = (
        pd.to_numeric(
            base.get(
                "Manual_Total_Brazos_min",
                pd.Series(index=base.index, dtype=float),
            ),
            errors="coerce",
        )
    )

    mask_binario = auto_min.notna() & manual_min.notna()
    ciclos_binarios = int(mask_binario.sum())
    horas_auto = float(auto_min[mask_binario].sum()) / 60.0 if ciclos_binarios else 0.0

    horas_auto_txt = f"{horas_auto:,.2f} h".replace(",", "X").replace(".", ",").replace("X", ".")

    # ------------------------------------------------------
    # Sin operador registrado
    # ------------------------------------------------------
    if "Operador_Filtro" in base.columns:
        op = (
            base["Operador_Filtro"]
            .fillna("SIN DATO")
            .astype(str)
            .str.strip()
            .str.upper()
        )
        sin_operador = int(
            op.isin(["", "SIN DATO", "NONE", "NAN"]).sum()
        )
    else:
        campos_op = [
            c for c in ["Operador_ZDA", "Operador", "Operario"]
            if c in base.columns
        ]
        if campos_op:
            tiene_op = pd.Series(False, index=base.index)
            for campo in campos_op:
                vals = base[campo].fillna("").astype(str).str.strip()
                tiene_op = tiene_op | vals.ne("")
            sin_operador = int((~tiene_op).sum())
        else:
            sin_operador = total_ciclos

    pct_sin_operador = (
        sin_operador / total_ciclos * 100
        if total_ciclos > 0
        else 0.0
    )

    # ------------------------------------------------------
    # Tarjetas
    # ------------------------------------------------------
    st.markdown(
        """
        <style>
        .ebr-kpi-card {
            min-height: 142px;
            border: 1px solid #e2e8f0;
            border-radius: 18px;
            background: #ffffff;
            padding: 1.15rem 1.25rem 1.05rem 1.25rem;
            box-shadow: 0 4px 14px rgba(15, 23, 42, 0.06);
            margin-bottom: 0.45rem;
        }
        .ebr-kpi-label {
            font-size: 0.78rem;
            letter-spacing: 0.055em;
            text-transform: uppercase;
            color: #8a8a84;
            font-weight: 500;
            margin-bottom: 0.50rem;
        }
        .ebr-kpi-value {
            font-size: 2.00rem;
            line-height: 1.05;
            color: #111111;
            font-weight: 750;
            margin-bottom: 0.48rem;
        }
        .ebr-kpi-sub {
            font-size: 0.95rem;
            line-height: 1.30;
            color: #66645f;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    cards = [
        (
            "Ciclos cargados",
            f"{total_ciclos}",
            jumbo_detalle,
        ),
        (
            "Rango de fechas",
            rango_fecha,
            fechas_sub,
        ),
        (
            "Horas automático (total)",
            horas_auto_txt,
            f"{ciclos_binarios} ciclos con dato automático/manual",
        ),
        (
            "Sin operador registrado",
            f"{sin_operador}",
            f"{pct_sin_operador:.0f}% de los ciclos",
        ),
    ]

    cols = st.columns(4)
    for col, (label, value, sub) in zip(cols, cards):
        with col:
            st.markdown(
                f"""
                <div class="ebr-kpi-card">
                    <div class="ebr-kpi-label">{label}</div>
                    <div class="ebr-kpi-value">{value}</div>
                    <div class="ebr-kpi-sub">{sub}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )



# ==========================================================
# BLOQUE 1 - AUTOMATIZACIÓN
# ==========================================================

@fragment
def render_automation_section(
    df_automatico: pd.DataFrame,
    sel_jumbos,
    sel_tipos,
    sel_rocas,
    sel_operadores,
    mostrar_auto: bool,
    mostrar_linea_auto: bool,
    mostrar_arm: bool,
):
    if df_automatico.empty:
        st.info("Sin datos suficientes de automatización.")
        return

    render_kpis_uso_automatico(df_automatico)
    st.markdown("<div style='height:0.35rem'></div>", unsafe_allow_html=True)

    df_visible = df_automatico[
        df_automatico["Jumbo"].astype(str).isin(
            [str(x) for x in sel_jumbos]
        )
        & df_automatico["Tipo_Disparo"].isin(sel_tipos)
        & df_automatico["Tipo_Roca"].isin(sel_rocas)
        & df_automatico["Operador_Filtro"].isin(sel_operadores)
    ].copy()

    # El rango de fechas del sidebar también filtra los gráficos de
    # Uso Automático para que sean consistentes con las tarjetas superiores.
    df_visible = aplicar_filtro_fechas_global(df_visible)

    st.subheader("Evolución del movimiento automático")

    fig_auto = grafico_auto(
        df_visible,
        mostrar_auto,
        mostrar_linea_auto,
    )

    if fig_auto is not None:
        st.plotly_chart(
            fig_auto,
            width="stretch",
            config={"displaylogo": False},
        )
    else:
        st.info(
            "No hay ciclos visibles con datos de movimiento automático "
            "para los filtros globales seleccionados."
        )

    st.subheader("Evolución del movimiento automático por operador")
    st.caption(
        "Cada color representa un operador y cada punto un ciclo/round. "
        "El porcentaje global de la leyenda se calcula con la suma de los "
        "tiempos automáticos y manuales de los ciclos visibles. "
        "El símbolo del punto identifica el jumbo: círculo = JUMB001, "
        "cuadrado = JUMB002."
    )

    fig_auto_operador = grafico_auto_por_operador(
        df_visible,
        mostrar_auto,
        mostrar_linea_auto,
    )

    if fig_auto_operador is not None:
        st.plotly_chart(
            fig_auto_operador,
            width="stretch",
            config={"displaylogo": False},
        )
    else:
        st.info(
            "No hay ciclos visibles con operador identificado y datos de "
            "movimiento automático para los filtros seleccionados."
        )

    st.subheader("Uso automático por brazo")

    jumbos_visibles = sorted(
        df_visible.get(
            "Jumbo",
            pd.Series(dtype=object),
        ).dropna().astype(str).unique()
    )

    if not jumbos_visibles:
        st.info(
            "No hay jumbos visibles con los filtros globales seleccionados."
        )
        return

    hubo_grafico = False

    for jumbo in jumbos_visibles:
        fig_arm = grafico_brazos(
            df_visible,
            jumbo,
            mostrar_arm,
        )
        if fig_arm is not None:
            hubo_grafico = True
            st.plotly_chart(
                fig_arm,
                width="stretch",
                config={"displaylogo": False},
            )

    if not hubo_grafico:
        st.info(
            "No hay datos por brazo para los filtros globales seleccionados."
        )


# ==========================================================
# BLOQUE 2 - BARRENOS CUT
# ==========================================================

@fragment
def render_cut_section(
    df_resumen: pd.DataFrame,
    df_reportes: pd.DataFrame,
    sel_jumbos,
    sel_tipos,
    sel_rocas,
    sel_operadores,
    mostrar_cut: bool,
):
    st.subheader("Evolución de la longitud perforada en barrenos Cut")

    df_cut = preparar_cut(
        df_resumen,
        df_reportes,
    )

    if df_cut.empty:
        st.info("Sin datos suficientes de barrenos Cut.")
        return

    st.caption(
        "Cada punto representa la mediana de la longitud perforada "
        "de los barrenos Cut de cada ciclo. Se aplican los filtros "
        "globales del panel lateral."
    )

    fig_cut = grafico_cut(
        df_cut,
        sel_jumbos,
        sel_tipos,
        sel_rocas,
        sel_operadores,
        mostrar_cut,
    )

    if fig_cut is not None:
        st.plotly_chart(
            fig_cut,
            width="stretch",
            config={"displaylogo": False},
        )
    else:
        st.info(
            "No hay ciclos visibles con los filtros globales seleccionados."
        )


# ==========================================================
# BLOQUE 3 - TIEMPOS DE CICLO DE PERFORACIÓN
# ==========================================================

@fragment
def render_zda_section(
    df_zda: pd.DataFrame,
    sel_jumbos,
    sel_tipos,
    sel_rocas,
    sel_operadores,
    mostrar_zda: bool,
):
    st.subheader("Tiempos de ciclo de perforación")

    if (
        df_zda.empty
        or not {
            "Inicio_Perforacion_TS",
            "Fin_Perforacion_TS",
        }.issubset(df_zda.columns)
    ):
        st.info(
            "Sin datos ZDA suficientes para mostrar tiempos de ciclo."
        )
        return

    zda_all = df_zda[
        df_zda["Inicio_Perforacion_TS"].notna()
        & df_zda["Fin_Perforacion_TS"].notna()
    ].copy()

    if zda_all.empty:
        st.info("Sin ventanas de perforación ZDA válidas.")
        return

    st.caption(
        "Análisis de las ventanas reales de perforación por fecha operativa y turno. "
        "Turno día: 07:00–19:00 · Turno noche: 19:00–07:00."
    )

    # Clasificación unificada: Bottom + Easer + Cut + Contour.
    zda_all["Tipo_Disparo"] = (
        zda_all["Barrenos_Realizados"]
        .apply(clasificar_tipo_disparo_v33)
    )

    zda_rows = zda_all[
        zda_all["Jumbo"].astype(str).isin(
            [str(x) for x in sel_jumbos]
        )
        & zda_all["Tipo_Disparo"].isin(sel_tipos)
        & zda_all["Tipo_Roca"].isin(sel_rocas)
        & zda_all["Operador_Filtro"].isin(sel_operadores)
    ].copy()

    # ------------------------------------------------------
    # FILTRO GLOBAL DE FECHAS DEL PANEL LATERAL.
    # Los widgets se muestran siempre en el sidebar; aquí únicamente
    # se aplican los valores seleccionados a Tiempos de Ciclo.
    # ------------------------------------------------------
    fecha_inicio_zda = st.session_state.get("fecha_inicio_zda_global")
    fecha_fin_zda = st.session_state.get("fecha_fin_zda_global")

    if not zda_rows.empty:
        zda_rows["_Fecha_Operativa_Filtro"] = (
            zda_rows["Inicio_Perforacion_TS"]
            .apply(zda_operational_date)
        )

        fecha_op_date = pd.to_datetime(
            zda_rows["_Fecha_Operativa_Filtro"],
            errors="coerce",
            utc=True,
        ).dt.date

        if fecha_inicio_zda is not None and fecha_fin_zda is not None:
            if fecha_inicio_zda > fecha_fin_zda:
                st.error("La Fecha inicio no puede ser posterior a la Fecha fin.")
                zda_rows = zda_rows.iloc[0:0].copy()
            else:
                mask_fecha_zda = (
                    (fecha_op_date >= fecha_inicio_zda)
                    & (fecha_op_date <= fecha_fin_zda)
                )
                zda_rows = zda_rows[mask_fecha_zda].copy()

    st.caption(
        "Los filtros globales del panel lateral actualizan la gráfica "
        "y los resúmenes de esta sección."
    )

    st.markdown("#### Hora promedio de inicio por jumbo y turno")

    shift = resumen_turnos_zda(
        zda_rows
    )

    if not shift.empty:
        st.dataframe(
            shift,
            width="stretch",
            hide_index=True,
        )
        st.caption(
            "“Ciclos día/noche” cuenta todos los rounds del turno. "
            "Para promedio, inicio más temprano e inicio más tarde "
            "se considera solo el primer round de cada jumbo por "
            "fecha operativa y turno."
        )
    else:
        st.info(
            "No hay ciclos visibles para calcular los indicadores de inicio."
        )

    st.markdown("#### Hora promedio de fin por jumbo y turno")

    shift_fin = resumen_fin_turnos_zda(
        zda_rows
    )

    if not shift_fin.empty:
        st.dataframe(
            shift_fin,
            width="stretch",
            hide_index=True,
        )
        st.caption(
            "“Ciclos día/noche” cuenta todos los rounds iniciados en el turno. "
            "Para promedio, fin más temprano y fin más tarde se considera "
            "únicamente el término del último round de cada jumbo por "
            "fecha operativa y turno."
        )
    else:
        st.info(
            "No hay ciclos visibles para calcular los indicadores de fin."
        )

    with st.expander(
        "Tipo de disparo",
        expanded=False,
    ):
        type_summary = resumen_tipos_zda(
            zda_rows
        )

        if not type_summary.empty:
            st.dataframe(
                type_summary,
                width="stretch",
                hide_index=True,
            )
        else:
            st.info(
                "No hay ciclos visibles para resumir el tipo de disparo."
            )

        st.caption(
            "Criterio sobre barrenos de frente "
            "(Bottom + Easer + Cut + Contour). "
            "Reaming y Casing no se consideran para la clasificación."
        )


    ciclos_turno = preparar_timeline_ciclos_turno(
        zda_rows
    )

    if ciclos_turno.empty:
        st.markdown(
            "#### Piloto · Timeline de ciclos por día y turno"
        )
        st.info(
            "No hay ciclos suficientes para construir el timeline por turno."
        )
    else:
        st.markdown(
            "#### Piloto · Timeline de ciclos por día y turno"
        )
        st.caption(
            "Cada segmento representa un ciclo/round físico completo desde "
            "Inicio hasta Fin. JUMB001 se representa como contorno cerrado y "
            "JUMB002 como sólido, manteniendo la misma codificación en barras "
            "y puntos. El filtro de fechas de la cabecera aplica a ambos turnos."
        )

        ciclos_turno_filtrado = ciclos_turno.copy()

        solo_puntos_inicio_timeline = st.checkbox(
            "Mostrar solo puntos de inicio",
            value=False,
            key="solo_puntos_inicio_timeline",
            help=(
                "Oculta las barras completas y muestra únicamente la hora "
                "de inicio de cada ciclo para facilitar la identificación "
                "de patrones de arranque."
            ),
        )

        solo_primer_inicio_timeline = st.checkbox(
            "Mostrar solo el primer inicio por equipo y turno",
            value=False,
            key="solo_primer_inicio_timeline",
            disabled=not solo_puntos_inicio_timeline,
            help=(
                "Disponible cuando se activa la vista de puntos. "
                "Si un jumbo tiene dos o más ciclos en el mismo turno, "
                "solo se muestra el primer inicio y se ocultan los demás."
            ),
        )

        # Orden de lectura fijo: primero Turno Día y luego Turno Noche.
        fig_turno_dia = grafico_timeline_ciclos_turno(
            ciclos_turno_filtrado,
            "Día",
            solo_puntos_inicio=solo_puntos_inicio_timeline,
            solo_primer_inicio=solo_primer_inicio_timeline,
        )

        if fig_turno_dia is not None:
            st.plotly_chart(
                fig_turno_dia,
                width="stretch",
                config={
                    "displaylogo": False,
                    "scrollZoom": False,
                },
            )
        else:
            st.info(
                "No hay ciclos visibles en el Turno Día."
            )

        fig_turno_noche = grafico_timeline_ciclos_turno(
            ciclos_turno_filtrado,
            "Noche",
            solo_puntos_inicio=solo_puntos_inicio_timeline,
            solo_primer_inicio=solo_primer_inicio_timeline,
        )

        if fig_turno_noche is not None:
            st.plotly_chart(
                fig_turno_noche,
                width="stretch",
                config={
                    "displaylogo": False,
                    "scrollZoom": False,
                },
            )
        else:
            st.info(
                "No hay ciclos visibles en el Turno Noche."
            )

        # ------------------------------------------------------
        # DISTRIBUCIÓN DE PRIMEROS INICIOS
        # ------------------------------------------------------
        st.markdown(
            "#### Distribución de primeros inicios"
        )
        st.caption(
            "El histograma agrupa el primer inicio de cada jumbo por fecha y turno "
            "en intervalos de 30 minutos. La curva suavizada permite identificar "
            "visualmente las horas de mayor concentración."
        )

        primeros_distribucion = preparar_primeros_inicios_distribucion(
            ciclos_turno_filtrado
        )

        if primeros_distribucion.empty:
            st.info(
                "No hay información suficiente para construir la distribución "
                "de primeros inicios."
            )
        else:
            fig_dist_dia = grafico_distribucion_primeros_inicios(
                primeros_distribucion,
                "Día",
            )
            if fig_dist_dia is not None:
                st.plotly_chart(
                    fig_dist_dia,
                    width="stretch",
                    config={
                        "displaylogo": False,
                        "scrollZoom": False,
                    },
                )

            fig_dist_noche = grafico_distribucion_primeros_inicios(
                primeros_distribucion,
                "Noche",
            )
            if fig_dist_noche is not None:
                st.plotly_chart(
                    fig_dist_noche,
                    width="stretch",
                    config={
                        "displaylogo": False,
                        "scrollZoom": False,
                    },
                )

        # ------------------------------------------------------
        # TENDENCIA DIARIA DEL PRIMER INICIO CONSOLIDADO
        # ------------------------------------------------------
        st.markdown(
            "#### Tendencia diaria de la hora de primer inicio"
        )
        st.caption(
            "Para cada fecha y turno se toma únicamente el primer inicio "
            "de cada jumbo disponible. Luego se calcula el promedio y la "
            "mediana entre los equipos. Con 2 equipos ambos valores son "
            "iguales; con 3 o más pueden diferenciarse."
        )

        ctrl_prom, ctrl_med, _ = st.columns([1.0, 1.0, 3.2])
        with ctrl_prom:
            mostrar_promedio_inicio = st.checkbox(
                "Mostrar promedio",
                value=True,
                key="mostrar_promedio_inicio_diario",
            )
        with ctrl_med:
            mostrar_mediana_inicio = st.checkbox(
                "Mostrar mediana",
                value=True,
                key="mostrar_mediana_inicio_diario",
            )

        resumen_inicio_diario = preparar_tendencia_inicio_diario(
            ciclos_turno_filtrado
        )

        if not mostrar_promedio_inicio and not mostrar_mediana_inicio:
            st.info(
                "Activa Promedio y/o Mediana para mostrar la tendencia."
            )
        elif resumen_inicio_diario.empty:
            st.info(
                "No hay información suficiente para calcular la tendencia diaria."
            )
        else:
            fig_inicio_dia = grafico_tendencia_inicio_diario(
                resumen_inicio_diario,
                "Día",
                mostrar_promedio=mostrar_promedio_inicio,
                mostrar_mediana=mostrar_mediana_inicio,
            )
            if fig_inicio_dia is not None:
                st.plotly_chart(
                    fig_inicio_dia,
                    width="stretch",
                    config={
                        "displaylogo": False,
                        "scrollZoom": False,
                    },
                )

            fig_inicio_noche = grafico_tendencia_inicio_diario(
                resumen_inicio_diario,
                "Noche",
                mostrar_promedio=mostrar_promedio_inicio,
                mostrar_mediana=mostrar_mediana_inicio,
            )
            if fig_inicio_noche is not None:
                st.plotly_chart(
                    fig_inicio_noche,
                    width="stretch",
                    config={
                        "displaylogo": False,
                        "scrollZoom": False,
                    },
                )

        # ------------------------------------------------------
        # TENDENCIA DIARIA DEL TÉRMINO DEL ÚLTIMO CICLO
        # ------------------------------------------------------
        st.markdown(
            "#### Tendencia diaria de la hora de término del último ciclo"
        )
        st.caption(
            "Para cada fecha y turno se toma el Fin más tardío de cada jumbo. "
            "Si un equipo realizó dos o más rounds, se considera el término "
            "del último round. Luego se calcula el promedio y la mediana "
            "entre los equipos disponibles."
        )

        ctrl_prom_fin, ctrl_med_fin, _ = st.columns([1.0, 1.0, 3.2])
        with ctrl_prom_fin:
            mostrar_promedio_fin = st.checkbox(
                "Mostrar promedio",
                value=True,
                key="mostrar_promedio_ultimo_fin_diario",
            )
        with ctrl_med_fin:
            mostrar_mediana_fin = st.checkbox(
                "Mostrar mediana",
                value=True,
                key="mostrar_mediana_ultimo_fin_diario",
            )

        resumen_ultimo_fin_diario = preparar_tendencia_ultimo_fin_diario(
            ciclos_turno_filtrado
        )

        if not mostrar_promedio_fin and not mostrar_mediana_fin:
            st.info(
                "Activa Promedio y/o Mediana para mostrar la tendencia."
            )
        elif resumen_ultimo_fin_diario.empty:
            st.info(
                "No hay información suficiente para calcular el término "
                "diario del último ciclo."
            )
        else:
            fig_fin_dia = grafico_tendencia_ultimo_fin_diario(
                resumen_ultimo_fin_diario,
                "Día",
                mostrar_promedio=mostrar_promedio_fin,
                mostrar_mediana=mostrar_mediana_fin,
            )
            if fig_fin_dia is not None:
                st.plotly_chart(
                    fig_fin_dia,
                    width="stretch",
                    config={
                        "displaylogo": False,
                        "scrollZoom": False,
                    },
                )

            fig_fin_noche = grafico_tendencia_ultimo_fin_diario(
                resumen_ultimo_fin_diario,
                "Noche",
                mostrar_promedio=mostrar_promedio_fin,
                mostrar_mediana=mostrar_mediana_fin,
            )
            if fig_fin_noche is not None:
                st.plotly_chart(
                    fig_fin_noche,
                    width="stretch",
                    config={
                        "displaylogo": False,
                        "scrollZoom": False,
                    },
                )

        with st.expander(
            "Detalle de ciclos mostrados en el timeline",
            expanded=False,
        ):
            detalle_timeline = ciclos_turno_filtrado[
                [
                    "Fecha_Operativa",
                    "Turno",
                    "Jumbo",
                    "Ciclo",
                    "Inicio",
                    "Fin",
                    "Duracion",
                    "Tipo_Disparo",
                    "Tipo_Roca",
                    "Labor",
                    "Operador",
                    "Barrenos",
                    "Sobrepasa_Turno",
                ]
            ].copy()

            st.dataframe(
                detalle_timeline,
                width="stretch",
                hide_index=True,
            )


# ==========================================================
# BLOQUE 4 - CLASIFICACIÓN DE DISPAROS
# ==========================================================

@fragment
def render_classification_section(
    df_reportes: pd.DataFrame,
    df_automatico: pd.DataFrame,
    df_atipicos: pd.DataFrame,
    sel_jumbos,
    sel_tipos,
    sel_rocas,
    sel_operadores,
):
    filtrados = df_reportes[
        df_reportes["Jumbo"].astype(str).isin(
            [str(x) for x in sel_jumbos]
        )
        & df_reportes["Tipo_Disparo"].isin(sel_tipos)
        & df_reportes["Tipo_Roca"].isin(sel_rocas)
        & df_reportes["Operador_Filtro"].isin(sel_operadores)
    ].copy()

    st.caption(
        "Esta sección utiliza los mismos filtros globales del panel lateral."
    )

    col_class, col_read = st.columns(
        [1, 1.25]
    )

    with col_class:
        if filtrados.empty:
            st.info(
                "No hay reportes visibles con los filtros globales seleccionados."
            )
        else:
            resumen_clase = (
                filtrados["Tipo_Disparo"]
                .value_counts()
                .reindex(
                    TIPOS_DISPARO,
                    fill_value=0,
                )
                .rename_axis(
                    "Tipo de disparo"
                )
                .reset_index(
                    name="N reportes"
                )
            )

            resumen_clase = resumen_clase[
                resumen_clase["N reportes"] > 0
            ]

            st.dataframe(
                resumen_clase,
                width="stretch",
                hide_index=True,
            )

    with col_read:
        st.markdown("#### Resumen de lectura")

        pdf_rows = (
            filtrados[
                filtrados["Fuente"].eq("PDF")
            ]
            if "Fuente" in filtrados.columns
            else pd.DataFrame()
        )

        zda_rows_all = (
            filtrados[
                filtrados["Fuente"].eq("ZDA")
            ]
            if "Fuente" in filtrados.columns
            else pd.DataFrame()
        )

        lectura_ok = int(
            filtrados.get(
                "Lectura_Confiable",
                pd.Series(dtype=object),
            ).eq("OK").sum()
        )

        pdf_ok = (
            int(
                pdf_rows.get(
                    "Lectura_Confiable",
                    pd.Series(dtype=object),
                ).eq("OK").sum()
            )
            if not pdf_rows.empty
            else 0
        )

        zda_ok = (
            int(
                zda_rows_all.get(
                    "Lectura_Confiable",
                    pd.Series(dtype=object),
                ).eq("OK").sum()
            )
            if not zda_rows_all.empty
            else 0
        )

        conteo_ok = int(
            filtrados.get(
                "Estado_Conteo",
                pd.Series(dtype=object),
            ).eq("OK").sum()
        )

        metros_ok = int(
            filtrados.get(
                "Estado_Metros_Tipos",
                pd.Series(dtype=object),
            ).eq("OK").sum()
        )

        # Atípicos pertenecientes a ciclos actualmente visibles.
        n_atipicos_visible = 0
        if (
            not df_atipicos.empty
            and {"Jumbo", "Ciclo"}.issubset(df_atipicos.columns)
            and {"Jumbo", "Ciclo"}.issubset(filtrados.columns)
        ):
            claves_visibles = {
                (str(j), str(c))
                for j, c in zip(
                    filtrados["Jumbo"],
                    filtrados["Ciclo"],
                )
            }

            n_atipicos_visible = sum(
                (str(j), str(c)) in claves_visibles
                for j, c in zip(
                    df_atipicos["Jumbo"],
                    df_atipicos["Ciclo"],
                )
            )

        r1, r2, r3, r4, r5 = st.columns(5)

        r1.metric(
            "Archivos",
            len(filtrados),
        )
        r2.metric(
            "PDF",
            len(pdf_rows),
        )
        r3.metric(
            "ZDA",
            len(zda_rows_all),
        )
        r4.metric(
            "Lectura PDF OK",
            f"{pdf_ok}/{len(pdf_rows)}",
        )
        r5.metric(
            "Lectura ZDA OK",
            f"{zda_ok}/{len(zda_rows_all)}",
        )

        s1, s2, s3, s4 = st.columns(4)

        s1.metric(
            "Conteo OK",
            f"{conteo_ok}/{len(filtrados)}",
        )
        s2.metric(
            "Metros por tipo OK",
            f"{metros_ok}/{len(filtrados)}",
        )
        s3.metric(
            "Revisar lectura",
            len(filtrados) - lectura_ok,
        )
        s4.metric(
            "Atípicos",
            n_atipicos_visible,
        )

    st.markdown(
        "#### Uso automático por ciclo"
    )

    auto_filtrado = df_automatico[
        df_automatico["Jumbo"].astype(str).isin(
            [str(x) for x in sel_jumbos]
        )
        & df_automatico["Tipo_Disparo"].isin(sel_tipos)
        & df_automatico["Tipo_Roca"].isin(sel_rocas)
        & df_automatico["Operador_Filtro"].isin(sel_operadores)
    ].copy()

    cols_auto = [
        c
        for c in [
            "Fecha_Inicio",
            "Jumbo",
            "Ciclo",
            "Operador_Filtro",
            "Barrenos_Realizados",
            "Tipo_Disparo",
            "Considerado_KPI_Automatizacion",
            "Auto_Total_Brazos_min",
            "Manual_Total_Brazos_min",
            "Pct_Movimiento_Automatico_Brazos",
            "Pct_Automatico_Brazo1",
            "Pct_Automatico_Brazo2",
            "Gap_Automatico_Brazos_pp",
        ]
        if c in auto_filtrado.columns
    ]

    tabla_auto = auto_filtrado[
        cols_auto
    ].copy()

    tabla_auto = tabla_auto.rename(
        columns={
            "Operador_Filtro": "Operador",
            "Considerado_KPI_Automatizacion": "KPI Auto",
            "Auto_Total_Brazos_min": "Auto total min",
            "Manual_Total_Brazos_min": "Manual total min",
            "Pct_Movimiento_Automatico_Brazos": "Auto Jumbo %",
            "Pct_Automatico_Brazo1": "Brazo 1 %",
            "Pct_Automatico_Brazo2": "Brazo 2 %",
            "Gap_Automatico_Brazos_pp": "Gap brazos pp",
        }
    )

    if "KPI Auto" in tabla_auto.columns:
        tabla_auto["KPI Auto"] = (
            tabla_auto["KPI Auto"]
            .map(
                {
                    True: "Sí",
                    False: "No",
                }
            )
        )

    st.dataframe(
        tabla_auto,
        width="stretch",
        hide_index=True,
        height=360,
    )


# ==========================================================
# BLOQUE 5 - RESULTADOS POR ARCHIVO
# ==========================================================

@fragment
def render_resultados_section(resultados_validos):
    st.caption(
        f"{len(resultados_validos)} archivo(s) procesado(s) acumulado(s)"
    )

    # Un expander cerrado igualmente ejecuta/renderiza su contenido en Streamlit.
    # Para no serializar decenas de imágenes/tablas a la vez, se muestran
    # 10 ciclos por vista. La navegación se hace por ciclo, no por tamaño de página.
    n_total = len(resultados_validos)
    page_size = 10
    n_pages = max(1, (n_total + page_size - 1) // page_size)

    # Mantener la página actual dentro del rango válido.
    page_key = "resultados_page_nav"
    if page_key not in st.session_state:
        st.session_state[page_key] = 1
    st.session_state[page_key] = max(
        1,
        min(int(st.session_state[page_key]), n_pages),
    )

    # Selector directo: permite escribir/buscar por jumbo, ciclo o fecha.
    cycle_indices = list(range(n_total))

    def _cycle_label(i):
        rr = resultados_validos[i].get("resumen_reporte", {})
        fuente_i = rr.get("Fuente") or resultados_validos[i].get("fuente") or ""
        return (
            f"{rr.get('Jumbo', '-')} · "
            f"Ciclo {rr.get('Ciclo', '-')} · "
            f"{rr.get('Fecha_Inicio', '-')}"
            + (f" · {fuente_i}" if fuente_i else "")
        )

    jump_key = "resultados_jump_cycle"

    def _jump_to_cycle():
        selected_idx = st.session_state.get(jump_key)
        if selected_idx is not None:
            st.session_state[page_key] = int(selected_idx) // page_size + 1

    st.selectbox(
        "Ir directamente a un ciclo",
        options=cycle_indices,
        index=None,
        placeholder="Buscar por jumbo, ciclo o fecha...",
        format_func=_cycle_label,
        key=jump_key,
        on_change=_jump_to_cycle,
    )

    page = int(st.session_state[page_key])

    nav_prev, nav_info, nav_next = st.columns([1.1, 3.0, 1.1])

    with nav_prev:
        if st.button(
            "← Anterior",
            key="resultados_prev",
            disabled=page <= 1,
            width="stretch",
        ):
            page = max(1, page - 1)
            st.session_state[page_key] = page

    with nav_next:
        if st.button(
            "Siguiente →",
            key="resultados_next",
            disabled=page >= n_pages,
            width="stretch",
        ):
            page = min(n_pages, page + 1)
            st.session_state[page_key] = page

    start = (page - 1) * page_size
    end = min(start + page_size, n_total)

    with nav_info:
        st.markdown(
            f"<div style='text-align:center; padding-top:0.45rem;'>"
            f"<strong>Ciclos {start + 1}–{end} de {n_total}</strong>"
            f"<br><span style='color:#667085; font-size:0.86rem;'>"
            f"Página {page} de {n_pages}</span></div>",
            unsafe_allow_html=True,
        )

    desglosar = st.checkbox(
        "Desglosar todos los ciclos de esta página",
        value=False,
        key="desglosar_todos",
    )

    for idx, r in enumerate(resultados_validos[start:end], start=start):
        rep = r["resumen_reporte"]
        fuente = (
            rep.get("Fuente")
            or r.get("fuente")
            or "PDF"
        )
        titulo = (
            f"{rep.get('Jumbo','-')} · "
            f"Ciclo {rep.get('Ciclo','-')} · "
            f"{rep.get('Fecha_Inicio','-')} · "
            f"{fuente}"
        )

        with st.expander(
            titulo,
            expanded=desglosar,
        ):
            col_info, col_nav = st.columns([4.7, 1.3])

            with col_info:
                m1, m2, m3, m4, m5, m6, m7 = st.columns(7)

                m1.metric(
                    "Serie",
                    rep.get("Numero_Serie") or "-",
                )
                m2.metric(
                    "Operador",
                    rep.get("Operador_ZDA")
                    or rep.get("Operador")
                    or rep.get("Operario")
                    or "-",
                    help=(
                        rep.get("Fuente_Operador")
                        or ("PDF · campo Operario" if rep.get("Operario") else None)
                    ),
                )
                m3.metric(
                    "Sección",
                    seccion_desde_plan_texto(
                        rep.get("Plan_Perforacion")
                    ),
                )
                m4.metric(
                    "Tipo de roca",
                    rep.get("Tipo_Roca")
                    or tipo_roca_desde_plan_texto(
                        rep.get("Plan_Perforacion")
                    ),
                )
                m5.metric(
                    "Tipo de disparo",
                    rep.get("Tipo_Disparo") or "-",
                )
                m6.metric(
                    "Barrenos",
                    int(rep["Barrenos_Realizados"])
                    if pd.notna(rep.get("Barrenos_Realizados"))
                    else "-",
                )
                m7.metric(
                    "Metros perforados",
                    fmt(
                        rep.get("Metros_Perforados"),
                        2,
                        " m",
                    ),
                )

                a0, a1, a2, a3, a4 = st.columns(5)
                a0.metric(
                    "Lectura",
                    rep.get("Lectura_Confiable")
                    or rep.get("Estado")
                    or "-",
                )
                a1.metric(
                    "Movimiento automático",
                    fmt(
                        rep.get(
                            "Pct_Movimiento_Automatico_Brazos"
                        ),
                        1,
                        "%",
                    ),
                )
                a2.metric(
                    "Movimiento manual",
                    fmt(
                        rep.get(
                            "Pct_Movimiento_Manual_Brazos"
                        ),
                        1,
                        "%",
                    ),
                )
                a3.metric(
                    "Brazo 1 automático",
                    fmt(
                        rep.get("Pct_Automatico_Brazo1"),
                        1,
                        "%",
                    ),
                )
                a4.metric(
                    "Brazo 2 automático",
                    fmt(
                        rep.get("Pct_Automatico_Brazo2"),
                        1,
                        "%",
                    ),
                )

                # Conteo individual por brazo para este ciclo.
                detalle_round = r.get("detalle")
                b1_count = b2_count = 0
                if (
                    isinstance(detalle_round, pd.DataFrame)
                    and not detalle_round.empty
                    and "Boom" in detalle_round.columns
                ):
                    boom_round = pd.to_numeric(
                        detalle_round["Boom"],
                        errors="coerce",
                    )
                    b1_count = int(boom_round.eq(1).sum())
                    b2_count = int(boom_round.eq(2).sum())

                b1m, b2m, btm = st.columns(3)
                b1m.metric("Barrenos Brazo 1", b1_count)
                b2m.metric("Barrenos Brazo 2", b2_count)
                btm.metric("Total por brazos", b1_count + b2_count)

                if fuente == "ZDA":
                    z1, z2, z3 = st.columns(3)
                    z1.metric(
                        "Inicio perforación real",
                        rep.get("Inicio_Perforacion") or "-",
                    )
                    z2.metric(
                        "Fin perforación real",
                        rep.get("Fin_Perforacion") or "-",
                    )
                    z3.metric(
                        "Tiempo de perforación",
                        rep.get("Tiempo_Perforacion_hms") or "-",
                    )

            detalle_key = f"detalle_ciclo_{r.get('_cache_key', idx)}"
            mostrar_detalle = st.toggle(
                "Cargar gráficos y plano de este ciclo",
                value=False,
                key=detalle_key,
                help="Los visuales se generan bajo demanda para mantener bajo el uso de memoria.",
            )

            box_path = nav_path = None
            if mostrar_detalle:
                with st.spinner("Generando/cargando visuales del ciclo..."):
                    box_path, nav_path = asegurar_visuales_resultado(r)

            with col_nav:
                if nav_path:
                    if fuente == "ZDA":
                        st.caption(
                            "Plano reconstruido desde ZDA · "
                            f"sección {seccion_desde_plan_texto(rep.get('Plan_Perforacion'))}"
                        )
                    else:
                        st.caption("Plano de navegación del PDF")
                    st.image(str(nav_path), width="stretch")
                elif not mostrar_detalle:
                    st.caption("Plano disponible bajo demanda")

            if box_path:
                st.image(str(box_path), width="stretch")
                with open(box_path, "rb") as fh:
                    png_bytes = fh.read()
                st.download_button(
                    "Descargar gráfico PNG",
                    png_bytes,
                    file_name=(
                        f"{rep.get('Jumbo','JUMBO')}_"
                        f"Ciclo_{rep.get('Ciclo','-')}.png"
                    ),
                    mime="image/png",
                    key=f"png_{idx}",
                )
                del png_bytes

            val = r.get("validacion")
            if (
                isinstance(val, pd.DataFrame)
                and not val.empty
            ):
                with st.expander(
                    "Validación",
                    expanded=False,
                ):
                    cols = [
                        c
                        for c in [
                            "Tipo",
                            "Esperado",
                            "Encontrado",
                            "Diferencia",
                            "Estado",
                        ]
                        if c in val.columns
                    ]
                    st.dataframe(
                        val[cols],
                        width="stretch",
                        hide_index=True,
                        height=220,
                    )

            vm = r.get("validacion_metros")
            if (
                isinstance(vm, pd.DataFrame)
                and not vm.empty
            ):
                with st.expander(
                    "Metros por tipo",
                    expanded=False,
                ):
                    cols = [
                        c
                        for c in [
                            "Tipo",
                            "N",
                            "Metros_Reporte_m",
                            "Metros_Extraidos_m",
                            "Diferencia_m",
                            "Estado",
                        ]
                        if c in vm.columns
                    ]
                    st.dataframe(
                        vm[cols],
                        width="stretch",
                        hide_index=True,
                        height=220,
                    )

            extras = r.get("extras")
            if (
                isinstance(extras, pd.DataFrame)
                and not extras.empty
            ):
                st.markdown(
                    f"#### Barrenos extra ({len(extras)})"
                )
                cols = [
                    c
                    for c in [
                        "ID",
                        "Tipo",
                        "Longitud_roca_m",
                        "Beta_grados",
                    ]
                    if c in extras.columns
                ]
                st.dataframe(
                    extras[cols],
                    width="stretch",
                    hide_index=True,
                    height=180,
                )

            mwd = r.get("mwd_barrenos")
            if (
                isinstance(mwd, pd.DataFrame)
                and not mwd.empty
            ):
                with st.expander(
                    f"MWD por brazo y secuencia ({len(mwd)})",
                    expanded=False,
                ):
                    cols = [
                        c
                        for c in [
                            "Brazo",
                            "Secuencia",
                            "Estado_MWD",
                            "Profundidad_Max_MWD_m",
                            "Muestras_MWD",
                            "Inicio_MWD",
                            "Fin_MWD",
                            "Duracion_MWD_s",
                        ]
                        if c in mwd.columns
                    ]
                    st.dataframe(
                        mwd[cols],
                        width="stretch",
                        hide_index=True,
                        height=260,
                    )


# Si ningún operador quedó seleccionado/detectado, no bloquear los gráficos.
if not global_operadores and "Operador_Filtro" in df_reportes.columns:
    global_operadores = sorted(
        df_reportes["Operador_Filtro"]
        .fillna("SIN DATO")
        .astype(str)
        .unique()
        .tolist()
    )


# ==========================================================
# PRESENTACIÓN POR SECCIONES
# ==========================================================

st.divider()
st.header("Consolidado")
st.markdown(
    """
    <div style="
        padding: 0.65rem 0 0.25rem 0;
        font-size: 1.05rem;
        font-weight: 700;
        color: #1f2937;
    ">
        Secciones del análisis
    </div>
    """,
    unsafe_allow_html=True,
)
st.caption(
    "Selecciona una sección para mostrar únicamente ese grupo de análisis. "
    "Se resaltan como accesos principales del app."
)

SECCIONES_ANALISIS = [
    "Uso Automático",
    "Perforación",
    "Tiempos de Ciclo",
    "Clasificación",
    "Resultados por archivo",
]

if "seccion_analisis_principal" not in st.session_state:
    st.session_state["seccion_analisis_principal"] = SECCIONES_ANALISIS[0]


def _cambiar_seccion_analisis(seccion):
    """
    Callback de navegación.

    Streamlit ejecuta el callback antes del rerun completo, por lo que
    la nueva sección ya está guardada cuando se vuelven a dibujar los
    botones. Así el resaltado cambia con el primer clic.
    """
    st.session_state["seccion_analisis_principal"] = seccion


st.markdown(
    """
    <style>
    div[data-testid="stHorizontalBlock"] div[data-testid="column"] .stButton > button.seccion-nav {
        min-height: 68px;
        font-size: 1.02rem;
        font-weight: 700;
        border-radius: 14px;
        border: 1.5px solid #d1d5db;
        box-shadow: 0 1px 2px rgba(0,0,0,0.04);
        white-space: normal;
        line-height: 1.15;
        padding: 0.70rem 0.75rem;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

cols_sec = st.columns(len(SECCIONES_ANALISIS))
for i, seccion in enumerate(SECCIONES_ANALISIS):
    activa = st.session_state.get("seccion_analisis_principal") == seccion
    cols_sec[i].button(
        seccion,
        key=f"btn_seccion_{i}",
        width="stretch",
        type="primary" if activa else "secondary",
        help=f"Ir a la sección: {seccion}",
        on_click=_cambiar_seccion_analisis,
        args=(seccion,),
    )

seccion_activa = st.session_state.get(
    "seccion_analisis_principal",
    SECCIONES_ANALISIS[0],
)

st.markdown(
    """
    <script>
    const marcarBotonesSeccion = () => {
      const bloques = window.parent.document.querySelectorAll('button[kind]');
      bloques.forEach((btn) => {
        const txt = (btn.innerText || "").trim();
        if (
          txt.includes("Uso Automático") ||
          txt.includes("Perforación") ||
          txt.includes("Tiempos de Ciclo") ||
          txt.includes("Clasificación") ||
          txt.includes("Resultados por archivo")
        ) {
          btn.classList.add("seccion-nav");
        }
      });
    };
    setTimeout(marcarBotonesSeccion, 100);
    setTimeout(marcarBotonesSeccion, 600);
    </script>
    """,
    unsafe_allow_html=True,
)

if seccion_activa == "Uso Automático":
    with st.container(border=True):
        render_automation_section(
            df_automatico,
            global_jumbos,
            global_tipos,
            global_rocas,
            global_operadores,
            global_lbl_auto,
            global_line_auto,
            global_lbl_arm,
        )

elif seccion_activa == "Perforación":
    with st.container(border=True):
        render_cut_section(
            df_resumen,
            df_reportes,
            global_jumbos,
            global_tipos,
            global_rocas,
            global_operadores,
            global_lbl_cut,
        )

elif seccion_activa == "Tiempos de Ciclo":
    with st.container(border=True):
        render_zda_section(
            df_zda,
            global_jumbos,
            global_tipos,
            global_rocas,
            global_operadores,
            global_lbl_zda,
        )

elif seccion_activa == "Clasificación":
    with st.container(border=True):
        st.subheader("Clasificación")
        render_classification_section(
            df_reportes,
            df_automatico,
            df_atipicos,
            global_jumbos,
            global_tipos,
            global_rocas,
            global_operadores,
        )

elif seccion_activa == "Resultados por archivo":
    with st.container(border=True):
        st.subheader("Resultados por archivo")
        render_resultados_section(
            resultados_validos,
        )

        if errores:
            st.divider()
            st.subheader("Archivos con error")
            st.dataframe(
                pd.DataFrame([
                    {
                        "Archivo": r.get("nombre_archivo"),
                        "Error": r.get("error"),
                    }
                    for r in errores
                ]),
                width="stretch",
                hide_index=True,
            )
